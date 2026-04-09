import streamlit as st
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
from data_loader import load_config, save_config, get_all_reports, fetch_notion_data, load_product_list, fetch_instructor_info
import json
import os
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import zipfile
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email.mime.text import MIMEText
from email import encoders

st.set_page_config(page_title="TRNT 레슨리포트", page_icon="🏋️", layout="wide")

# ── 비밀번호 인증 ──
if "authenticated" not in st.session_state:
    st.session_state["authenticated"] = False

if not st.session_state["authenticated"]:
    st.markdown("## 🏋️ TRNT 필라테스 레슨리포트")
    pw = st.text_input("비밀번호를 입력하세요", type="password", key="login_pw")
    if st.button("로그인", type="primary"):
        cfg = json.load(open(os.path.join(os.path.dirname(__file__), "config.json"), "r"))
        if pw == cfg.get("dashboard_password", "trnt1234"):
            st.session_state["authenticated"] = True
            st.rerun()
        else:
            st.error("비밀번호가 올바르지 않습니다.")
    st.stop()

st.markdown("""
<style>
    .main .block-container { padding-top: 1rem; max-width: 1400px; }

    /* 메트릭 카드 */
    .metric-card {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 1rem; border-radius: 12px; color: white; text-align: center;
        box-shadow: 0 4px 15px rgba(0,0,0,0.1); margin-bottom: 0.5rem;
        min-height: 140px; display: flex; flex-direction: column; justify-content: center;
    }
    .metric-card.green { background: linear-gradient(135deg, #11998e 0%, #38ef7d 100%); }
    .metric-card.orange { background: linear-gradient(135deg, #F2994A 0%, #F2C94C 100%); }
    .metric-card.blue { background: linear-gradient(135deg, #2193b0 0%, #6dd5ed 100%); }
    .metric-card.red { background: linear-gradient(135deg, #e53935 0%, #ef5350 100%); }
    .metric-card.gray { background: linear-gradient(135deg, #636e72 0%, #b2bec3 100%); }
    .metric-card .value { font-size: 1.8rem; font-weight: 700; margin: 0.2rem 0; white-space: nowrap; }
    .metric-card .label { font-size: 0.75rem; opacity: 0.9; white-space: nowrap; overflow: hidden; text-overflow: ellipsis; }
    .metric-card .sub { font-size: 0.75rem; opacity: 0.8; margin-top: 2px; }

    /* 섹션 헤더 */
    .section-header {
        background: #f8f9fa; border-left: 4px solid #2C3E50;
        padding: 0.6rem 1rem; margin: 1.5rem 0 1rem 0;
        font-size: 1.1rem; font-weight: 700; color: #2C3E50;
        border-radius: 0 8px 8px 0;
    }
    .section-header.personal { border-left-color: #3498db; }
    .section-header.duet { border-left-color: #9b59b6; }
    .section-header.group { border-left-color: #27ae60; }
    .section-header.trial { border-left-color: #f39c12; }
    .section-header.renewal { border-left-color: #e74c3c; }

    /* 코멘트 박스 */
    .comment-box {
        padding: 1rem 1.2rem; border-radius: 10px; margin: 0.8rem 0;
        font-size: 0.95rem; line-height: 1.5;
    }
    .comment-box.good {
        background: #d4edda; border: 1px solid #c3e6cb; color: #155724;
    }
    .comment-box.warn {
        background: #fff3cd; border: 1px solid #ffeeba; color: #856404;
    }
    .comment-box.danger {
        background: #f8d7da; border: 1px solid #f5c6cb; color: #721c24;
    }

    /* MVP 카드 */
    .mvp-card {
        background: linear-gradient(135deg, #FFD700 0%, #FFA500 100%);
        padding: 1.2rem; border-radius: 14px; color: #333; text-align: center;
        box-shadow: 0 6px 20px rgba(255,165,0,0.3);
        min-height: 140px; display: flex; flex-direction: column; justify-content: center;
    }
    .mvp-card .title { font-size: 0.9rem; font-weight: 600; }
    .mvp-card .name { font-size: 1.4rem; font-weight: 800; margin: 0.3rem 0; }
    .mvp-card .detail { font-size: 0.8rem; opacity: 0.85; }

    /* 3개월 비교 테이블 */
    .compare-table {
        width: 100%; border-collapse: collapse; margin: 0.5rem 0;
        font-size: 0.9rem;
    }
    .compare-table th {
        background: #2C3E50; color: white; padding: 8px 12px;
        text-align: center; font-weight: 600;
    }
    .compare-table td {
        padding: 7px 12px; text-align: center; border-bottom: 1px solid #eee;
    }
    .compare-table tr:nth-child(even) { background: #f8f9fa; }
    .compare-table .highlight { font-weight: 700; color: #2C3E50; }
    .compare-table .up { color: #27ae60; }
    .compare-table .down { color: #e74c3c; }

    /* 핵심지표 카드 (크게) */
    .key-metric {
        background: #fff; border: 2px solid #2C3E50; border-radius: 14px;
        padding: 1.2rem; text-align: center; margin-bottom: 0.5rem;
        box-shadow: 0 2px 8px rgba(0,0,0,0.06);
        min-height: 180px; display: flex; flex-direction: column; justify-content: center;
    }
    .key-metric .km-label { font-size: 0.85rem; color: #666; font-weight: 600; }
    .key-metric .km-value { font-size: 2.2rem; font-weight: 800; margin: 0.3rem 0; }
    .key-metric .km-formula { font-size: 0.7rem; color: #999; }
    .key-metric .km-sub { font-size: 0.78rem; color: #888; margin-top: 4px; }
    .key-metric.good { border-color: #27ae60; }
    .key-metric.good .km-value { color: #27ae60; }
    .key-metric.bad { border-color: #e74c3c; }
    .key-metric.bad .km-value { color: #e74c3c; }
    .key-metric.warn { border-color: #f39c12; }
    .key-metric.warn .km-value { color: #f39c12; }
    .key-metric.neutral .km-value { color: #2C3E50; }

    /* 팀원 테이블 */
    .member-table {
        width: 100%; border-collapse: collapse; margin: 0.5rem 0; font-size: 0.88rem;
    }
    .member-table th {
        background: #34495e; color: white; padding: 10px 14px;
        text-align: center; font-weight: 600; font-size: 0.85rem;
    }
    .member-table td {
        padding: 9px 14px; text-align: center; border-bottom: 1px solid #eee;
    }
    .member-table tr:hover { background: #eef2f7; }
    .member-table .good-val { color: #27ae60; font-weight: 700; }
    .member-table .bad-val { color: #e74c3c; font-weight: 700; }
    .member-table .warn-val { color: #f39c12; font-weight: 700; }

    div[data-testid="stSidebar"] { background-color: #1a1a2e; }
    div[data-testid="stSidebar"] .stMarkdown { color: #e0e0e0; }
    .stTabs [data-baseweb="tab-list"] { gap: 8px; }
    .stTabs [data-baseweb="tab"] {
        background-color: #f0f2f6; border-radius: 8px 8px 0 0;
        padding: 8px 20px; font-weight: 600;
    }
    .stTabs [aria-selected="true"] { background-color: #2C3E50; color: white; }
</style>
""", unsafe_allow_html=True)


def metric_card(label, value, sub=None, color=""):
    sub_html = f'<div class="sub">{sub}</div>' if sub else ""
    st.markdown(f'''
        <div class="metric-card {color}">
            <div class="label">{label}</div>
            <div class="value">{value}</div>
            {sub_html}
        </div>
    ''', unsafe_allow_html=True)


def key_metric_card(label, value, formula="", sub="", target=None, as_pct=True):
    """핵심 4대 지표용 큰 카드. as_pct=False면 소수점 표시"""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        css_class = "neutral"
        display = "-"
    else:
        display = f"{value:.1%}" if as_pct else f"{value:.2f}"
        if target is not None:
            css_class = "good" if value >= target else ("warn" if value >= target * 0.8 else "bad")
        else:
            css_class = "neutral"
    parts = [f'<div class="key-metric {css_class}">',
             f'<div class="km-label">{label}</div>',
             f'<div class="km-value">{display}</div>']
    if formula:
        parts.append(f'<div class="km-formula">{formula}</div>')
    if sub:
        parts.append(f'<div class="km-sub">{sub}</div>')
    parts.append('</div>')
    st.markdown("".join(parts), unsafe_allow_html=True)


def fmt_pct(v):
    if v is None or pd.isna(v):
        return "-"
    return f"{v:.1%}"


def fmt_rate(v):
    """개인출석률, 그룹출석율 등 소수점 표시 (% 없이)"""
    if v is None or pd.isna(v):
        return "-"
    return f"{v:.2f}"


def fmt_num(v):
    if v is None or pd.isna(v):
        return "-"
    return str(int(round(v)))


def delta_str(curr, prev):
    """이전 대비 증감 문자열"""
    if curr is None or prev is None or pd.isna(curr) or pd.isna(prev):
        return ""
    diff = curr - prev
    if abs(diff) < 0.001:
        return "→"
    arrow = "▲" if diff > 0 else "▼"
    if isinstance(curr, float) and abs(curr) <= 2:  # percentage
        return f"{arrow} {abs(diff):.1%}p"
    return f"{arrow} {abs(diff):.0f}"


def get_comment(r, config):
    """강사 성과 기반 격려 코멘트 생성"""
    targets = config.get("targets", {})
    comments = []
    level = "good"

    # 개인출석달성율 체크
    achieve = r.get("개인출석달성율")
    if achieve is not None and not pd.isna(achieve):
        if achieve >= 1.0:
            comments.append(f"🎉 개인레슨 목표 달성! ({fmt_pct(achieve)})")
        elif achieve >= 0.85:
            comments.append(f"💪 목표까지 거의 다 왔어요! ({fmt_pct(achieve)})")
            level = "warn"
        else:
            gap = r.get("목표레슨수", 0) or 0
            actual = r.get("개인레슨+OT", 0) or 0
            need = gap - actual
            if need > 0:
                comments.append(f"📌 목표까지 {need:.0f}건 남았어요. 화이팅!")
            else:
                comments.append(f"📌 개인출석달성율 {fmt_pct(achieve)} — 조금 더 힘내봐요!")
            level = "danger"

    # 재등록율
    re_rate = r.get("재등록율")
    target_re = targets.get("재등록률_목표", 0.7)
    if re_rate is not None and not pd.isna(re_rate):
        if re_rate >= target_re:
            comments.append(f"✅ 재등록율 {fmt_pct(re_rate)} — 훌륭해요!")
        else:
            comments.append(f"📋 재등록율 {fmt_pct(re_rate)} (목표 {fmt_pct(target_re)})")

    # 체험승률
    trial = r.get("체험승률")
    target_trial = targets.get("체험등록율_목표", 0.5)
    if trial is not None and not pd.isna(trial):
        if trial >= target_trial:
            comments.append(f"⭐ 체험승률 {fmt_pct(trial)} — 대단해요!")
        elif trial > 0:
            comments.append(f"🔔 체험승률 {fmt_pct(trial)} (목표 {fmt_pct(target_trial)})")

    if not comments:
        comments.append("📊 이번 달 데이터를 확인해보세요.")
        level = "warn"

    return level, " | ".join(comments)


def section_header(title, css_class=""):
    st.markdown(f'<div class="section-header {css_class}">{title}</div>', unsafe_allow_html=True)


def render_3month_table(rows, columns, pct_cols=None):
    """최근 3개월 비교 테이블 HTML"""
    if pct_cols is None:
        pct_cols = set()
    else:
        pct_cols = set(pct_cols)

    header = "<tr>" + "".join(f"<th>{c}</th>" for c in columns) + "</tr>"
    body = ""
    for row in rows:
        cells = ""
        for i, val in enumerate(row):
            col = columns[i]
            if col == "항목":
                cells += f'<td class="highlight">{val}</td>'
            elif val is None or (isinstance(val, float) and pd.isna(val)):
                cells += '<td>-</td>'
            elif col in pct_cols or (isinstance(val, float) and abs(val) <= 2 and col != "항목"):
                # check if it's a percentage column
                if col in pct_cols:
                    cells += f'<td>{val:.1%}</td>'
                else:
                    cells += f'<td>{val:.1f}</td>'
            elif isinstance(val, float):
                cells += f'<td>{val:.1f}</td>'
            elif isinstance(val, (int, float)):
                cells += f'<td>{val}</td>'
            else:
                cells += f'<td>{val}</td>'
        body += f"<tr>{cells}</tr>"

    return f'<table class="compare-table">{header}{body}</table>'


@st.cache_data(ttl=300)
def load_data():
    config = load_config()
    reports = get_all_reports(config)
    return config, reports


# Sidebar
with st.sidebar:
    st.markdown("## 🏋️ TRNT 필라테스")
    st.markdown("### 레슨리포트 대시보드")
    st.markdown("---")
    page = st.radio("메뉴", ["📊 대시보드", "👤 강사별 리포트", "📈 월별 비교", "👥 팀별 비교", "⚙️ 설정"],
                    label_visibility="collapsed")
    st.markdown("---")
    if st.button("🔄 데이터 새로고침"):
        st.cache_data.clear()
        st.rerun()
    st.markdown("---")
    if st.button("📥 노션 데이터 업데이트"):
        with st.spinner("노션에서 데이터 가져오는 중..."):
            try:
                cfg = load_config()
                notion = fetch_notion_data(cfg)
                notion_path = os.path.join(os.path.dirname(__file__), "notion_cache.json")
                with open(notion_path, "w") as f:
                    json.dump(notion, f, ensure_ascii=False, indent=2)
                st.cache_data.clear()
                st.success(f"노션 데이터 업데이트 완료! ({sum(len(v) for v in notion.values())}건)")
                st.rerun()
            except Exception as e:
                st.error(f"노션 연결 실패: {e}")

config, reports = load_data()

# 퇴사/휴직 필터링: staff에 등록되어 있고 status가 "퇴사" 또는 "휴직"인 강사 제외
_excluded = {nm for nm, info in config.get("staff", {}).items() if info.get("status") in ("퇴사", "휴직")}
if _excluded:
    for m in reports:
        reports[m] = reports[m][~reports[m]["강사"].isin(_excluded)].reset_index(drop=True)

months = sorted(reports.keys())

# 월 정렬 함수 (연도+월 숫자 기반)
def month_sort_key(m):
    # "2025년 3월" → (2025, 3), "3월" → (0, 3)
    m = m.strip()
    if "년" in m:
        parts = m.split("년")
        try:
            year = int(parts[0].strip())
        except ValueError:
            year = 0
        month_part = parts[1].replace("월", "").strip()
    else:
        year = 0
        month_part = m.replace("월", "").strip()
    try:
        month = int(month_part)
    except ValueError:
        month = 0
    return (year, month)

months = sorted(months, key=month_sort_key)


def get_instructor_row(instructor, month):
    """특정 월의 강사 데이터 행 반환"""
    if month not in reports:
        return None
    df = reports[month]
    row = df[df["강사"] == instructor]
    if row.empty:
        return None
    return row.iloc[0]


def get_recent_months(n=3):
    """최근 n개월 반환"""
    return months[-n:] if len(months) >= n else months


def generate_instructor_report(instructor, recent_months, reports_data, config):
    """강사별 개인 리포트 엑셀 생성"""
    wb = Workbook()
    ws = wb.active
    ws.title = "레슨리포트"

    # 스타일 정의
    title_font = Font(name="맑은 고딕", size=16, bold=True, color="FFFFFF")
    header_font = Font(name="맑은 고딕", size=11, bold=True, color="FFFFFF")
    section_font = Font(name="맑은 고딕", size=12, bold=True, color="2C3E50")
    data_font = Font(name="맑은 고딕", size=10)
    bold_font = Font(name="맑은 고딕", size=10, bold=True)
    pct_font_good = Font(name="맑은 고딕", size=10, bold=True, color="27AE60")
    pct_font_bad = Font(name="맑은 고딕", size=10, bold=True, color="E74C3C")
    title_fill = PatternFill("solid", fgColor="2C3E50")
    header_fill = PatternFill("solid", fgColor="3498DB")
    section_fill = PatternFill("solid", fgColor="ECF0F1")
    light_blue = PatternFill("solid", fgColor="EBF5FB")
    thin_border = Border(
        left=Side(style="thin", color="D5D8DC"),
        right=Side(style="thin", color="D5D8DC"),
        top=Side(style="thin", color="D5D8DC"),
        bottom=Side(style="thin", color="D5D8DC"),
    )
    center = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    # 열 너비 설정
    ws.column_dimensions["A"].width = 18
    for i, m in enumerate(recent_months):
        col = get_column_letter(i + 2)
        ws.column_dimensions[col].width = 16

    row = 1
    # 타이틀
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(recent_months) + 1)
    cell = ws.cell(row=row, column=1, value=f"TRNT 필라테스 — {instructor} 레슨리포트")
    cell.font = title_font
    cell.fill = title_fill
    cell.alignment = center
    ws.row_dimensions[row].height = 35
    row += 1

    # 기간 표시
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(recent_months) + 1)
    period = f"{recent_months[0]} ~ {recent_months[-1]}"
    cell = ws.cell(row=row, column=1, value=period)
    cell.font = Font(name="맑은 고딕", size=10, color="7F8C8D")
    cell.alignment = center
    row += 2

    def write_header(row_num):
        ws.cell(row=row_num, column=1, value="항목").font = header_font
        ws.cell(row=row_num, column=1).fill = header_fill
        ws.cell(row=row_num, column=1).alignment = center
        ws.cell(row=row_num, column=1).border = thin_border
        for i, m in enumerate(recent_months):
            c = ws.cell(row=row_num, column=i + 2, value=m)
            c.font = header_font
            c.fill = header_fill
            c.alignment = center
            c.border = thin_border
        return row_num + 1

    def write_section(row_num, title):
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=len(recent_months) + 1)
        cell = ws.cell(row=row_num, column=1, value=title)
        cell.font = section_font
        cell.fill = section_fill
        cell.alignment = left_align
        cell.border = thin_border
        ws.row_dimensions[row_num].height = 25
        return row_num + 1

    def write_row(row_num, label, values, fmt="num", highlight=False):
        c = ws.cell(row=row_num, column=1, value=label)
        c.font = bold_font
        c.border = thin_border
        c.alignment = left_align
        if highlight:
            c.fill = light_blue
        for i, v in enumerate(values):
            cell = ws.cell(row=row_num, column=i + 2)
            cell.border = thin_border
            cell.alignment = center
            if highlight:
                cell.fill = light_blue
            if v is None or (isinstance(v, float) and pd.isna(v)):
                cell.value = "-"
                cell.font = data_font
            elif fmt == "pct":
                cell.value = v
                cell.number_format = "0.0%"
                targets = config.get("targets", {})
                cell.font = data_font
            elif fmt == "rate":
                cell.value = round(v, 2)
                cell.number_format = "0.00"
                cell.font = data_font
            else:
                cell.value = int(round(v)) if isinstance(v, (int, float)) else v
                cell.font = data_font
        return row_num + 1

    # 데이터 수집
    month_data = []
    for m in recent_months:
        r = get_instructor_row(instructor, m)
        month_data.append(r)

    def vals(key):
        return [d[key] if d is not None else None for d in month_data]

    # ===== 핵심 4대 지표 =====
    row = write_section(row, "⭐ 핵심 4대 지표")
    row = write_header(row)
    row = write_row(row, "개인출석률", vals("개인출석률"), "rate", True)
    row = write_row(row, "그룹출석율", vals("그룹출석율"), "rate", True)
    row = write_row(row, "재등록율", vals("재등록율"), "pct", True)
    row = write_row(row, "체험승률", vals("체험승률"), "pct", True)
    row += 1

    # ===== 개인레슨 =====
    row = write_section(row, "🏃 개인레슨")
    row = write_header(row)
    row = write_row(row, "전체고객수", vals("전체고객수"))
    row = write_row(row, "홀딩고객수", vals("홀딩고객수"))
    row = write_row(row, "개인레슨수", vals("개인레슨수"))
    row = write_row(row, "개인OT수", vals("개인OT수"))
    row = write_row(row, "개인레슨+OT", vals("개인레슨+OT"), highlight=True)
    row = write_row(row, "목표레슨수", vals("목표레슨수"))
    row = write_row(row, "개인출석률", vals("개인출석률"), "rate")
    row = write_row(row, "출석달성율", vals("개인출석달성율"), "pct")
    row += 1

    # ===== 듀엣 =====
    row = write_section(row, "👥 듀엣")
    row = write_header(row)
    row = write_row(row, "듀엣회원수", vals("듀엣회원수"))
    row = write_row(row, "듀엣레슨수", vals("듀엣레슨수"))
    row = write_row(row, "듀엣출석률", vals("듀엣출석률"), "pct")
    row += 1

    # ===== 그룹 =====
    row = write_section(row, "🏢 그룹")
    row = write_header(row)
    row = write_row(row, "그룹회원수", vals("그룹회원수"))
    row = write_row(row, "그룹수업수", vals("그룹수업수"))
    row = write_row(row, "그룹출석수", vals("그룹출석수"))
    row = write_row(row, "그룹출석율", vals("그룹출석율"), "rate")
    row += 1

    # ===== 체험 =====
    row = write_section(row, "🌟 체험")
    row = write_header(row)
    row = write_row(row, "체험수업수", vals("체험수업수"))
    row = write_row(row, "체험등록수", vals("체험등록수"))
    row = write_row(row, "체험승률", vals("체험승률"), "pct")
    row += 1

    # ===== 재등록 =====
    row = write_section(row, "🔄 재등록")
    row = write_header(row)
    row = write_row(row, "재등예정수", vals("재등예정수"))
    row = write_row(row, "재등완료수", vals("재등완료수"))
    row = write_row(row, "재등록율", vals("재등록율"), "pct")
    row += 1

    # ===== 아카데미 =====
    row = write_section(row, "🎓 아카데미")
    row = write_header(row)
    row = write_row(row, "아카데미수업수", vals("아카데미수업수"))
    row = write_row(row, "딥코칭", vals("아카데미_딥코칭"))
    row = write_row(row, "모의테스트", vals("아카데미_모의테스트"))
    row = write_row(row, "체험", vals("아카데미_체험"))
    row = write_row(row, "그룹", vals("아카데미_그룹"))
    row += 1

    # ===== 종합 (스튜디오 평균 대비) =====
    row = write_section(row, "📊 종합 (스튜디오 평균 대비)")
    row = write_header(row)
    for metric_name, fmt_type in [("개인출석률", "rate"), ("그룹출석율", "rate"),
                                   ("재등록율", True), ("체험승률", True)]:
        my_vals = vals(metric_name)
        avg_vals = []
        for m in recent_months:
            if m in reports_data:
                col = reports_data[m][metric_name].dropna()
                avg_vals.append(col.mean() if len(col) > 0 else None)
            else:
                avg_vals.append(None)
        # 내 값 행
        row = write_row(row, metric_name, my_vals, highlight=True)
        row = write_row(row, f"  스튜디오 평균", avg_vals)

    # 인쇄 설정
    ws.print_area = f"A1:{get_column_letter(len(recent_months) + 1)}{row}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_all_reports_zip(all_instructors, recent_months, reports_data, config):
    """전체 강사 리포트를 ZIP으로 생성"""
    zip_buf = BytesIO()
    with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for inst in all_instructors:
            excel_buf = generate_instructor_report(inst, recent_months, reports_data, config)
            zf.writestr(f"{inst}_레슨리포트.xlsx", excel_buf.read())
    zip_buf.seek(0)
    return zip_buf


def build_report_html(instructor, recent_months, reports_data, config):
    """강사별 리포트 HTML 생성 (이메일용)"""
    month_data = []
    for m in recent_months:
        if m in reports_data:
            df = reports_data[m]
            row = df[df["강사"] == instructor]
            month_data.append(row.iloc[0] if not row.empty else None)
        else:
            month_data.append(None)

    r = month_data[-1]  # 최신 월
    current_month = recent_months[-1]
    targets = config.get("targets", {})

    def fv(val, fmt="num"):
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return "-"
        if fmt == "rate":
            return f"{val:.2f}"
        if fmt == "pct":
            return f"{val:.1%}"
        return str(int(round(val)))

    def card_color(val, target, is_rate=False):
        if val is None or (isinstance(val, float) and pd.isna(val)):
            return "#95a5a6"
        if is_rate:
            if val >= 1.5:
                return "#27ae60"
            if val >= 1.0:
                return "#f39c12"
            return "#e74c3c"
        if target and val >= target:
            return "#27ae60"
        if target and val >= target * 0.8:
            return "#f39c12"
        return "#e74c3c"

    def metric_card_html(label, value, color="#3498db", sub=""):
        sub_html = f'<div style="font-size:12px;color:#999;margin-top:4px;">{sub}</div>' if sub else ""
        return f'''<td style="width:25%;padding:8px;">
            <div style="background:{color};color:white;border-radius:12px;padding:16px;text-align:center;min-height:80px;">
                <div style="font-size:12px;opacity:0.9;">{label}</div>
                <div style="font-size:28px;font-weight:700;margin:6px 0;">{value}</div>
                {sub_html}
            </div></td>'''

    def key_card_html(label, value, color="#2C3E50", sub=""):
        sub_html = f'<div style="font-size:11px;color:#888;margin-top:4px;">{sub}</div>' if sub else ""
        return f'''<td style="width:25%;padding:8px;">
            <div style="border:2px solid {color};border-radius:12px;padding:16px;text-align:center;min-height:100px;">
                <div style="font-size:13px;color:#666;font-weight:600;">{label}</div>
                <div style="font-size:32px;font-weight:800;color:{color};margin:8px 0;">{value}</div>
                {sub_html}
            </div></td>'''

    def section_html(title):
        return f'<tr><td colspan="4" style="padding:20px 0 8px 0;"><div style="font-size:16px;font-weight:700;color:#2C3E50;border-left:4px solid #2C3E50;padding-left:10px;background:#f8f9fa;padding:8px 12px;">{title}</div></td></tr>'

    def table_header(cols):
        ths = "".join(f'<th style="background:#3498db;color:white;padding:10px;text-align:center;font-size:12px;">{c}</th>' for c in cols)
        return f'<tr>{ths}</tr>'

    def table_row(cells, highlight=False):
        bg = "background:#EBF5FB;" if highlight else ""
        tds = ""
        for i, c in enumerate(cells):
            fw = "font-weight:600;" if i == 0 else ""
            tds += f'<td style="padding:8px 10px;text-align:center;border-bottom:1px solid #eee;{fw}{bg}">{c}</td>'
        return f'<tr>{tds}</tr>'

    def vals(key, fmt="num"):
        return [fv(d[key], fmt) if d is not None else "-" for d in month_data]

    # ===== HTML 시작 =====
    html = f'''
    <div style="max-width:700px;margin:0 auto;font-family:'맑은 고딕',Arial,sans-serif;color:#333;">
        <div style="background:#2C3E50;color:white;padding:20px;text-align:center;border-radius:12px 12px 0 0;">
            <div style="font-size:14px;opacity:0.8;">TRNT 필라테스</div>
            <div style="font-size:24px;font-weight:700;margin:8px 0;">{instructor} 레슨리포트</div>
            <div style="font-size:13px;opacity:0.7;">{current_month}</div>
        </div>
        <div style="background:#fff;padding:20px;border:1px solid #eee;">
    '''

    if r is not None:
        # 핵심 4대 지표
        pa = r.get("개인출석률")
        ga = r.get("그룹출석율")
        rr = r.get("재등록율")
        tr = r.get("체험승률")

        pa_color = card_color(pa, 1.0, is_rate=True)
        ga_color = "#8e44ad" if ga and not pd.isna(ga) and ga >= 5.0 else "#3498db" if ga and not pd.isna(ga) and ga >= 4.0 else "#f39c12" if ga and not pd.isna(ga) and ga >= 3.0 else "#9b59b6" if ga and not pd.isna(ga) and ga >= 2.0 else "#e74c3c" if ga and not pd.isna(ga) else "#95a5a6"
        rr_color = card_color(rr, targets.get("재등록률_목표", 0.7))
        tr_color = card_color(tr, targets.get("체험등록율_목표", 0.5))

        html += '<table style="width:100%;border-collapse:collapse;"><tr>'
        html += key_card_html("개인출석률", fv(pa, "rate"), pa_color, f"레슨 {fv(r['개인레슨+OT'])} / 고객 {fv(r['전체고객수'])}")
        html += key_card_html("그룹출석율", fv(ga, "rate"), ga_color, f"출석 {fv(r['그룹출석수'])} / 수업 {fv(r['그룹수업수'])}")
        html += key_card_html("재등록율", fv(rr, "pct"), rr_color, f"완료 {fv(r['재등완료수'])} / 예정 {fv(r['재등예정수'])}")
        html += key_card_html("체험승률", fv(tr, "pct"), tr_color, f"등록 {fv(r['체험등록수'])} / 수업 {fv(r['체험수업수'])}")
        html += '</tr></table>'

        # 개인레슨
        html += '<table style="width:100%;border-collapse:collapse;">'
        html += section_html("개인레슨")
        html += '<tr>'
        html += metric_card_html("전체고객수", fv(r["전체고객수"]), "#2193b0")
        html += metric_card_html("홀딩고객수", fv(r["홀딩고객수"]), "#636e72")
        html += metric_card_html("개인레슨수", fv(r["개인레슨수"]), "#11998e")
        html += metric_card_html("개인OT", fv(r["개인OT수"]), "#11998e")
        html += '</tr><tr>'
        html += metric_card_html("개인레슨+OT", fv(r["개인레슨+OT"]), "#2193b0", f"목표: {fv(r['목표레슨수'])}")
        ach = r.get("개인출석달성율")
        ach_color = "#27ae60" if ach and not pd.isna(ach) and ach >= 1.0 else "#e74c3c"
        html += metric_card_html("개인출석률", fv(pa, "rate"), ach_color, f"달성율: {fv(ach, 'pct')}")
        html += '<td></td><td></td></tr>'
        html += '</table>'

        # 3개월 비교 테이블들
        if len(recent_months) > 1:
            sections = [
                ("개인레슨 추이", [
                    ("전체고객수", "num"), ("개인레슨수", "num"), ("개인OT수", "num"),
                    ("개인레슨+OT", "num"), ("개인출석률", "rate"), ("개인출석달성율", "pct"),
                ]),
                ("듀엣", [
                    ("듀엣회원수", "num"), ("듀엣레슨수", "num"), ("듀엣출석률", "pct"),
                ]),
                ("그룹", [
                    ("그룹회원수", "num"), ("그룹수업수", "num"), ("그룹출석수", "num"), ("그룹출석율", "rate"),
                ]),
                ("체험", [
                    ("체험수업수", "num"), ("체험등록수", "num"), ("체험승률", "pct"),
                ]),
                ("재등록", [
                    ("재등예정수", "num"), ("재등완료수", "num"), ("재등록율", "pct"),
                ]),
                ("아카데미", [
                    ("아카데미수업수", "num"), ("아카데미_딥코칭", "num"), ("아카데미_모의테스트", "num"),
                    ("아카데미_체험", "num"), ("아카데미_그룹", "num"),
                ]),
            ]
            for sec_title, items in sections:
                html += '<table style="width:100%;border-collapse:collapse;">'
                html += section_html(sec_title)
                html += '</table>'
                html += '<table style="width:100%;border-collapse:collapse;border:1px solid #eee;margin-bottom:12px;">'
                html += table_header(["항목"] + recent_months)
                for item_name, fmt in items:
                    label = item_name.replace("아카데미_", "")
                    hl = item_name in ("개인레슨+OT", "개인출석률", "그룹출석율", "재등록율", "체험승률")
                    html += table_row([label] + vals(item_name, fmt), highlight=hl)
                html += '</table>'

        # 종합 (스튜디오 평균 대비)
        html += '<table style="width:100%;border-collapse:collapse;">'
        html += section_html("종합 (스튜디오 평균 대비)")
        cur_m = recent_months[-1]
        avg_pa = reports_data[cur_m]["개인출석률"].dropna().mean() if cur_m in reports_data else None
        avg_ga = reports_data[cur_m]["그룹출석율"].dropna().mean() if cur_m in reports_data else None
        avg_rr = reports_data[cur_m]["재등록율"].dropna().mean() if cur_m in reports_data else None
        avg_tr = reports_data[cur_m]["체험승률"].dropna().mean() if cur_m in reports_data else None
        def _rate(v):
            return f"{v:.2f}" if v is not None and not pd.isna(v) else "-"
        def _pct(v):
            return f"{v:.1%}" if v is not None and not pd.isna(v) else "-"
        def _color(my, avg):
            if my is None or avg is None or pd.isna(my) or pd.isna(avg):
                return "#636e72"
            return "#27ae60" if my >= avg else "#e74c3c"
        html += '<tr>'
        html += metric_card_html("개인출석률", _rate(r["개인출석률"]),
                                  _color(r["개인출석률"], avg_pa), f"평균 {_rate(avg_pa)}")
        html += metric_card_html("그룹출석율", _rate(r["그룹출석율"]),
                                  _color(r["그룹출석율"], avg_ga), f"평균 {_rate(avg_ga)}")
        html += metric_card_html("재등록률", _pct(r["재등록율"]),
                                  _color(r["재등록율"], avg_rr), f"평균 {_pct(avg_rr)}")
        html += metric_card_html("체험승률", _pct(r["체험승률"]),
                                  _color(r["체험승률"], avg_tr), f"평균 {_pct(avg_tr)}")
        html += '</tr></table>'

    html += '''
        </div>
        <div style="background:#f8f9fa;padding:16px;text-align:center;border-radius:0 0 12px 12px;border:1px solid #eee;border-top:none;">
            <div style="font-size:12px;color:#999;">TRNT 필라테스 레슨리포트 자동 발송</div>
        </div>
    </div>'''
    return html


def send_report_email(instructor, email, recent_months, reports_data, config):
    """강사에게 레슨리포트 이메일 발송"""
    smtp_config = config.get("smtp", {})
    sender = smtp_config.get("sender_email", "")
    password = smtp_config.get("sender_password", "")
    smtp_server = smtp_config.get("server", "smtp.gmail.com")
    smtp_port = int(smtp_config.get("port", 587))

    if not sender or not password:
        return False, "SMTP 설정이 필요합니다. 설정 페이지에서 이메일 설정을 해주세요."

    current_month = recent_months[-1]
    html_body = build_report_html(instructor, recent_months, reports_data, config)
    excel_buf = generate_instructor_report(instructor, recent_months, reports_data, config)

    sender_name = smtp_config.get("sender_name", "TRNT 필라테스")
    msg = MIMEMultipart("alternative")
    msg["From"] = f"{sender_name} <{sender}>"
    msg["To"] = email
    msg["Subject"] = f"[TRNT] {current_month} 레슨리포트 — {instructor}"

    # 텍스트 버전 (HTML 못 보는 경우)
    msg.attach(MIMEText(f"{instructor} 선생님, {current_month} 레슨리포트입니다. HTML 이메일을 지원하는 클라이언트에서 확인해주세요.", "plain", "utf-8"))
    # HTML 버전
    msg.attach(MIMEText(html_body, "html", "utf-8"))

    # 엑셀 첨부
    part = MIMEBase("application", "vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    part.set_payload(excel_buf.read())
    encoders.encode_base64(part)
    part.add_header("Content-Disposition", "attachment",
                    filename=f"{instructor}_레슨리포트_{current_month}.xlsx")
    msg.attach(part)

    try:
        with smtplib.SMTP(smtp_server, smtp_port) as server:
            server.starttls()
            server.login(sender, password)
            server.send_message(msg)
        return True, "발송 완료"
    except Exception as e:
        return False, str(e)


# ============================================================
# PAGE: 대시보드
# ============================================================
if page == "📊 대시보드":
    st.markdown("# 📊 TRNT 필라테스 대시보드")

    if not months:
        st.warning("데이터가 없습니다. 설정에서 데이터 파일을 확인해주세요.")
        st.stop()

    selected_month = st.selectbox("월 선택", months, index=len(months) - 1)
    df = reports[selected_month]

    # ---- 이달의 MVP ----
    st.markdown("---")
    mvp_cols = st.columns(4)

    # 개인출석률 MVP
    df_personal = df.dropna(subset=["개인출석률"])
    if not df_personal.empty:
        top_personal = df_personal.loc[df_personal["개인출석률"].idxmax()]
        with mvp_cols[0]:
            st.markdown(f'''
            <div class="mvp-card">
                <div class="title">🏆 개인출석률 1위</div>
                <div class="name">{top_personal["강사"]}</div>
                <div class="detail">{fmt_rate(top_personal["개인출석률"])}</div>
            </div>''', unsafe_allow_html=True)

    # 그룹출석율 MVP
    df_group = df.dropna(subset=["그룹출석율"])
    if not df_group.empty:
        top_group = df_group.loc[df_group["그룹출석율"].idxmax()]
        with mvp_cols[1]:
            st.markdown(f'''
            <div class="mvp-card">
                <div class="title">🥇 그룹출석율 1위</div>
                <div class="name">{top_group["강사"]}</div>
                <div class="detail">{fmt_rate(top_group["그룹출석율"])}</div>
            </div>''', unsafe_allow_html=True)

    # 재등록율 MVP (동점시 재등완료수가 많은 강사 우선)
    df_re = df.dropna(subset=["재등록율"])
    if not df_re.empty:
        df_re_sorted = df_re.sort_values(["재등록율", "재등완료수"], ascending=[False, False])
        top_re = df_re_sorted.iloc[0]
        tied = df_re_sorted[df_re_sorted["재등록율"] == top_re["재등록율"]]
        tie_text = f" 외 {len(tied)-1}명" if len(tied) > 1 else ""
        with mvp_cols[2]:
            st.markdown(f'''
            <div class="mvp-card">
                <div class="title">🌟 재등록율 1위</div>
                <div class="name">{top_re["강사"]}{tie_text}</div>
                <div class="detail">{fmt_pct(top_re["재등록율"])} (완료 {fmt_num(top_re["재등완료수"])}건)</div>
            </div>''', unsafe_allow_html=True)

    # 체험승률 MVP (동점시 체험등록수가 많은 강사 우선)
    df_tr = df.dropna(subset=["체험승률"])
    if not df_tr.empty:
        df_tr_sorted = df_tr.sort_values(["체험승률", "체험등록수"], ascending=[False, False])
        top_tr = df_tr_sorted.iloc[0]
        tied_tr = df_tr_sorted[df_tr_sorted["체험승률"] == top_tr["체험승률"]]
        tie_text_tr = f" 외 {len(tied_tr)-1}명" if len(tied_tr) > 1 else ""
        with mvp_cols[3]:
            st.markdown(f'''
            <div class="mvp-card">
                <div class="title">⭐ 체험승률 1위</div>
                <div class="name">{top_tr["강사"]}{tie_text_tr}</div>
                <div class="detail">{fmt_pct(top_tr["체험승률"])} (등록 {fmt_num(top_tr["체험등록수"])}건)</div>
            </div>''', unsafe_allow_html=True)

    st.markdown("---")

    # ---- KPI 요약 (전체 평균 포함) ----
    targets = config.get("targets", {})
    c1, c2, c3, c4, c5, c6 = st.columns(6)
    with c1:
        metric_card("총 수업 건수", fmt_num(df["총수업수"].sum()),
                    sub=f"평균 {df['총수업수'].mean():.0f}건/인", color="blue")
    with c2:
        metric_card("개인레슨수", fmt_num(df["개인레슨수"].sum()),
                    sub=f"평균 {df['개인레슨수'].mean():.0f}건/인", color="green")
    with c3:
        avg_personal = df["개인출석률"].dropna().mean()
        t_pa = targets.get("목표레슨_배수", 2.0)
        pa_color = "green" if avg_personal and avg_personal >= t_pa else "orange" if avg_personal and avg_personal >= t_pa * 0.8 else "red"
        metric_card("개인출석률", fmt_rate(avg_personal), sub=f"목표 {t_pa:.1f}", color=pa_color)
    with c4:
        avg_group = df["그룹출석율"].dropna().mean()
        t_ga = targets.get("그룹출석율_목표", 5.0)
        ga_color = "green" if avg_group and avg_group >= t_ga else "orange" if avg_group and avg_group >= t_ga * 0.8 else "red"
        metric_card("그룹출석율", fmt_rate(avg_group), sub=f"목표 {t_ga:.1f}", color=ga_color)
    with c5:
        avg_trial = df["체험승률"].dropna().mean()
        t_tr = targets.get("체험등록율_목표", 0.50)
        tr_color = "green" if avg_trial and avg_trial >= t_tr else "orange" if avg_trial and avg_trial >= t_tr * 0.8 else "red"
        metric_card("평균 체험승률", fmt_pct(avg_trial), sub=f"목표 {t_tr:.0%}", color=tr_color)
    with c6:
        avg_re = df["재등록율"].dropna().mean()
        t_rr = targets.get("재등록률_목표", 0.70)
        rr_color = "green" if avg_re and avg_re >= t_rr else "orange" if avg_re and avg_re >= t_rr * 0.8 else "red"
        metric_card("평균 재등록율", fmt_pct(avg_re), sub=f"목표 {t_rr:.0%}", color=rr_color)

    st.markdown("---")

    # ---- 월별 추이 (전체 스튜디오) ----
    if len(months) > 1:
        st.markdown("### 📈 스튜디오 월별 추이")
        trend_data = []
        for m in months:
            mdf = reports[m]
            trend_data.append({
                "월": m,
                "총수업수": mdf["총수업수"].sum(),
                "개인레슨+OT": mdf["개인레슨+OT"].sum(),
                "평균개인출석률": mdf["개인출석률"].dropna().mean(),
                "평균그룹출석율": mdf["그룹출석율"].dropna().mean(),
                "평균재등록율": mdf["재등록율"].dropna().mean(),
                "평균체험승률": mdf["체험승률"].dropna().mean(),
            })
        trend_df = pd.DataFrame(trend_data)

        tcol1, tcol2 = st.columns(2)
        with tcol1:
            st.markdown("##### 출석률 추이 (소수점)")
            fig = go.Figure()
            fig.add_trace(go.Bar(x=trend_df["월"], y=trend_df["평균개인출석률"],
                                 name="개인출석률", marker_color="#3498db",
                                 text=[fmt_rate(v) for v in trend_df["평균개인출석률"]],
                                 textposition="outside"))
            fig.add_trace(go.Bar(x=trend_df["월"], y=trend_df["평균그룹출석율"],
                                 name="그룹출석율", marker_color="#27ae60",
                                 text=[fmt_rate(v) for v in trend_df["평균그룹출석율"]],
                                 textposition="outside"))
            fig.update_layout(height=300, margin=dict(l=0, r=0, t=30, b=0),
                             legend_title="", barmode="group")
            st.plotly_chart(fig, use_container_width=True)
        with tcol2:
            st.markdown("##### 등록율 추이 (%)")
            fig2 = go.Figure()
            fig2.add_trace(go.Scatter(x=trend_df["월"], y=trend_df["평균재등록율"],
                                      name="재등록율", mode="lines+markers+text",
                                      text=[fmt_pct(v) for v in trend_df["평균재등록율"]],
                                      textposition="top center", line=dict(color="#e74c3c", width=2)))
            fig2.add_trace(go.Scatter(x=trend_df["월"], y=trend_df["평균체험승률"],
                                      name="체험승률", mode="lines+markers+text",
                                      text=[fmt_pct(v) for v in trend_df["평균체험승률"]],
                                      textposition="bottom center", line=dict(color="#f39c12", width=2)))
            fig2.update_layout(height=300, margin=dict(l=0, r=0, t=30, b=0),
                              yaxis_tickformat=".0%", legend_title="")
            st.plotly_chart(fig2, use_container_width=True)

        st.markdown("---")

    # 핵심 4대 지표
    st.markdown("### ⭐ 핵심 4대 지표")

    # 개인출석률 / 그룹출석율 — 각각 가로막대 (높은순 정렬)
    col1, col2 = st.columns(2)
    with col1:
        st.markdown("#### 개인출석률")
        df_p = df.dropna(subset=["개인출석률"]).sort_values("개인출석률", ascending=True)
        if not df_p.empty:
            colors_p = ["#e74c3c" if v < 1.0 else "#f39c12" if v < 1.5 else "#27ae60" for v in df_p["개인출석률"]]
            fig_p = go.Figure(go.Bar(
                x=df_p["개인출석률"], y=df_p["강사"], orientation="h",
                marker_color=colors_p,
                text=[fmt_rate(v) for v in df_p["개인출석률"]], textposition="outside",
                hovertemplate="%{y}: %{text}<extra></extra>"
            ))
            fig_p.update_layout(height=500, margin=dict(l=0, r=0, t=10, b=0))
            st.plotly_chart(fig_p, use_container_width=True)

    with col2:
        st.markdown("#### 그룹출석율")
        df_g = df.dropna(subset=["그룹출석율"]).sort_values("그룹출석율", ascending=True)
        if not df_g.empty:
            colors_g = ["#e74c3c" if v < 2.0 else "#f39c12" if v < 3.0 else "#27ae60" if v < 4.0 else "#3498db" if v < 5.0 else "#8e44ad" for v in df_g["그룹출석율"]]
            fig_g = go.Figure(go.Bar(
                x=df_g["그룹출석율"], y=df_g["강사"], orientation="h",
                marker_color=colors_g,
                text=[fmt_rate(v) for v in df_g["그룹출석율"]], textposition="outside",
                hovertemplate="%{y}: %{text}<extra></extra>"
            ))
            fig_g.update_layout(height=500, margin=dict(l=0, r=0, t=10, b=0))
            st.plotly_chart(fig_g, use_container_width=True)

    # 재등록율 / 체험승률 — 꺾은선
    col3, col4 = st.columns(2)
    with col3:
        st.markdown("#### 강사별 재등록율")
        df_re = df.dropna(subset=["재등록율"]).sort_values("강사")
        if not df_re.empty:
            fig_re = go.Figure(go.Scatter(
                x=df_re["강사"], y=df_re["재등록율"], mode="lines+markers+text",
                line=dict(color="#e74c3c", width=2),
                text=[fmt_pct(v) for v in df_re["재등록율"]], textposition="top center"
            ))
            fig_re.update_layout(height=350, margin=dict(l=0, r=0, t=10, b=0),
                                  yaxis_tickformat=".0%", xaxis_tickangle=-45)
            st.plotly_chart(fig_re, use_container_width=True)
        else:
            st.info("재등록 데이터 없음")

    with col4:
        st.markdown("#### 강사별 체험승률")
        df_tr = df.dropna(subset=["체험승률"]).sort_values("강사")
        if not df_tr.empty:
            fig_tr = go.Figure(go.Scatter(
                x=df_tr["강사"], y=df_tr["체험승률"], mode="lines+markers+text",
                line=dict(color="#f39c12", width=2),
                text=[fmt_pct(v) for v in df_tr["체험승률"]], textposition="top center"
            ))
            fig_tr.update_layout(height=350, margin=dict(l=0, r=0, t=10, b=0),
                                  yaxis_tickformat=".0%", xaxis_tickangle=-45)
            st.plotly_chart(fig_tr, use_container_width=True)
        else:
            st.info("체험 데이터 없음")

    # 아카데미 현황
    if "아카데미_딥코칭" in df.columns:
        st.markdown("### 🎓 아카데미 현황")
        acad_data = df[["강사", "아카데미수업수", "아카데미_딥코칭", "아카데미_모의테스트", "아카데미_체험", "아카데미_그룹"]].copy()
        acad_data = acad_data[acad_data["아카데미수업수"] > 0].sort_values("아카데미수업수", ascending=False)
        if not acad_data.empty:
            acad_data.columns = ["강사", "전체", "딥코칭", "모의테스트", "체험", "그룹"]
            fig_acad = px.bar(acad_data, x="강사",
                              y=["딥코칭", "모의테스트", "체험", "그룹"],
                              barmode="stack", height=350,
                              color_discrete_sequence=["#3498db", "#f39c12", "#27ae60", "#9b59b6"])
            fig_acad.update_layout(margin=dict(l=0, r=0, t=10, b=0),
                                    legend_title="", xaxis_tickangle=-45)
            st.plotly_chart(fig_acad, use_container_width=True)
        else:
            st.info("아카데미 수업 데이터 없음")


# ============================================================
# PAGE: 강사별 리포트
# ============================================================
elif page == "👤 강사별 리포트":
    st.markdown("# 👤 강사별 리포트")

    if not months:
        st.warning("데이터가 없습니다.")
        st.stop()

    all_instructors = sorted(set().union(*[set(reports[m]["강사"]) for m in months]))

    ic1, ic2 = st.columns([2, 3])
    with ic1:
        selected_inst = st.selectbox("강사 선택", all_instructors)
    with ic2:
        month_mode = st.radio("비교 기간", ["최근 3개월", "전체", "직접 선택"], horizontal=True, key="month_mode")

    if month_mode == "최근 3개월":
        recent = get_recent_months(3)
    elif month_mode == "전체":
        recent = months[:]
    else:
        recent = st.multiselect("비교할 월 선택", months, default=get_recent_months(3), key="custom_months")
        if not recent:
            st.warning("1개 이상의 월을 선택해주세요.")
            st.stop()
        recent = sorted(recent, key=month_sort_key)

    current_month = recent[-1]
    r = get_instructor_row(selected_inst, current_month)

    if r is None:
        st.warning(f"{current_month} 데이터에 {selected_inst} 강사 정보가 없습니다.")
        st.stop()

    # ---- 격려 코멘트 ----
    level, comment_text = get_comment(r, config)
    st.markdown(f'<div class="comment-box {level}">{comment_text}</div>', unsafe_allow_html=True)

    # ---- 핵심 4대 지표 ----
    targets = config.get("targets", {})
    k1, k2, k3, k4 = st.columns(4)
    with k1:
        key_metric_card("개인출석률", r["개인출석률"],
                        sub=f"레슨 {fmt_num(r['개인레슨+OT'])} / 고객 {fmt_num(r['전체고객수'])}",
                        target=targets.get("개인출석율_목표", 0.85), as_pct=False)
    with k2:
        key_metric_card("그룹출석율", r["그룹출석율"],
                        sub=f"출석 {fmt_num(r['그룹출석수'])} / 수업 {fmt_num(r['그룹수업수'])}",
                        target=targets.get("그룹출석율_목표", 0.80), as_pct=False)
    with k3:
        key_metric_card("재등록율", r["재등록율"],
                        sub=f"완료 {fmt_num(r['재등완료수'])} / 예정 {fmt_num(r['재등예정수'])}",
                        target=targets.get("재등록률_목표", 0.70))
    with k4:
        key_metric_card("체험승률", r["체험승률"],
                        sub=f"등록 {fmt_num(r['체험등록수'])} / 수업 {fmt_num(r['체험수업수'])}",
                        target=targets.get("체험등록율_목표", 0.50))

    # ---- 종합 그래프 (최근 3개월 추이) ----
    if len(recent) > 1:
        st.markdown("### 📈 최근 월별 추이")
        trend_rows = []
        for m in recent:
            row = get_instructor_row(selected_inst, m)
            if row is not None:
                trend_rows.append({
                    "월": m,
                    "개인레슨+OT": row["개인레슨+OT"],
                    "듀엣레슨수": row["듀엣레슨수"],
                    "그룹출석수": row["그룹출석수"],
                    "총수업수": row["총수업수"],
                })
        if trend_rows:
            trend_df = pd.DataFrame(trend_rows)
            tcol1, tcol2 = st.columns(2)
            with tcol1:
                fig = px.bar(trend_df, x="월",
                             y=["개인레슨+OT", "듀엣레슨수", "그룹출석수"],
                             barmode="stack", height=280,
                             color_discrete_sequence=["#3498db", "#9b59b6", "#27ae60"])
                fig.update_layout(legend_title="", margin=dict(l=0, r=0, t=10, b=0))
                st.plotly_chart(fig, use_container_width=True)
            with tcol2:
                # 비율 추이
                rate_rows = []
                for m in recent:
                    row = get_instructor_row(selected_inst, m)
                    if row is not None:
                        rate_rows.append({"월": m,
                                          "개인출석률": row["개인출석률"],
                                          "그룹출석율": row["그룹출석율"],
                                          "재등록율": row["재등록율"],
                                          "체험승률": row["체험승률"]})
                if rate_rows:
                    rate_df = pd.DataFrame(rate_rows)
                    fig2 = go.Figure()
                    # 출석률 = 소수점 (좌측 Y축)
                    for col_name, color in [("개인출석률", "#3498db"), ("그룹출석율", "#27ae60")]:
                        vals = rate_df[col_name].tolist()
                        fig2.add_trace(go.Scatter(
                            x=rate_df["월"], y=vals, name=col_name,
                            mode="lines+markers+text",
                            text=[fmt_rate(v) if v is not None and not pd.isna(v) else "-" for v in vals],
                            textposition="top center", line=dict(color=color, width=2)
                        ))
                    # 등록율 = 퍼센트 (우측 Y축)
                    for col_name, color in [("재등록율", "#e74c3c"), ("체험승률", "#f39c12")]:
                        vals = rate_df[col_name].tolist()
                        fig2.add_trace(go.Scatter(
                            x=rate_df["월"], y=vals, name=col_name,
                            mode="lines+markers+text", yaxis="y2",
                            text=[fmt_pct(v) if v is not None and not pd.isna(v) else "-" for v in vals],
                            textposition="top center", line=dict(color=color, width=2, dash="dot")
                        ))
                    fig2.update_layout(height=280, margin=dict(l=0, r=0, t=10, b=0),
                                      yaxis=dict(title="출석률"),
                                      yaxis2=dict(title="등록율", overlaying="y", side="right", tickformat=".0%"),
                                      legend_title="")
                    st.plotly_chart(fig2, use_container_width=True)

    # ===== 개인레슨 섹션 =====
    section_header("🏃 개인레슨", "personal")

    pcol1, pcol2 = st.columns([2, 3])
    with pcol1:
        c1, c2 = st.columns(2)
        with c1:
            metric_card("전체고객수", fmt_num(r["전체고객수"]), color="blue")
        with c2:
            metric_card("홀딩고객수", fmt_num(r["홀딩고객수"]), color="gray")
        c3, c4 = st.columns(2)
        with c3:
            metric_card("개인레슨", fmt_num(r["개인레슨수"]), color="green")
        with c4:
            metric_card("개인OT", fmt_num(r["개인OT수"]), color="green")
        c5, c6 = st.columns(2)
        with c5:
            metric_card("개인레슨+OT", fmt_num(r["개인레슨+OT"]),
                        sub=f"목표: {fmt_num(r['목표레슨수'])}", color="blue")
        with c6:
            color = "green" if r["개인출석률"] and not pd.isna(r["개인출석률"]) and r["개인출석률"] >= 1.0 else "orange" if r["개인출석률"] and not pd.isna(r["개인출석률"]) and r["개인출석률"] >= 0.7 else "red"
            metric_card("개인출석률", fmt_rate(r["개인출석률"]),
                        sub=f"달성율: {fmt_pct(r['개인출석달성율'])}", color=color)

    with pcol2:
        if r["최소레슨수"] and not pd.isna(r["최소레슨수"]):
            actual = r["개인레슨+OT"]
            min_l = r["최소레슨수"]
            target_l = r["목표레슨수"]
            max_l = r["최대레슨수"]
            achieve_pct = (actual / target_l * 100) if target_l else 0

            # 프로그레스 바 시각화
            fig = go.Figure()

            # 배경 구간 (최소 / 목표 / 최대)
            fig.add_trace(go.Bar(
                y=["레슨수"], x=[min_l], name=f"최소 ({int(round(min_l))}건)",
                orientation="h", marker_color="#fadbd8",
                text=[f"최소 {int(round(min_l))}"], textposition="inside",
                textfont=dict(color="#999", size=11),
            ))
            fig.add_trace(go.Bar(
                y=["레슨수"], x=[target_l - min_l], name=f"목표 ({int(round(target_l))}건)",
                orientation="h", marker_color="#fdebd0",
                text=[f"목표 {int(round(target_l))}"], textposition="inside",
                textfont=dict(color="#999", size=11),
            ))
            fig.add_trace(go.Bar(
                y=["레슨수"], x=[max_l - target_l], name=f"최대 ({int(round(max_l))}건)",
                orientation="h", marker_color="#d5f5e3",
                text=[f"최대 {int(round(max_l))}"], textposition="inside",
                textfont=dict(color="#999", size=11),
            ))

            # 실제 레슨수 마커
            bar_color = "#27ae60" if actual >= target_l else "#f39c12" if actual >= min_l else "#e74c3c"
            fig.add_vline(x=actual, line_width=4, line_color=bar_color,
                          annotation_text=f"실제 {int(round(actual))}건 ({achieve_pct:.0f}%)",
                          annotation_position="top", annotation_font_size=13,
                          annotation_font_color=bar_color)

            fig.update_layout(
                barmode="stack", height=140,
                margin=dict(l=0, r=0, t=35, b=0),
                showlegend=True, legend=dict(orientation="h", yanchor="bottom", y=-0.6, font_size=11),
                yaxis=dict(visible=False),
                xaxis=dict(title="", range=[0, max_l * 1.1]),
            )
            st.plotly_chart(fig, use_container_width=True)

            # 달성 상태 한 줄 요약
            if actual >= target_l:
                st.markdown(f'<div class="comment-box good">🎉 목표 달성! 실제 {int(round(actual))}건 / 목표 {int(round(target_l))}건 (달성율 {achieve_pct:.0f}%)</div>', unsafe_allow_html=True)
            elif actual >= min_l:
                gap = int(round(target_l - actual))
                st.markdown(f'<div class="comment-box warn">💪 목표까지 {gap}건 남았어요! ({int(round(actual))}/{int(round(target_l))}건)</div>', unsafe_allow_html=True)
            else:
                gap = int(round(min_l - actual))
                st.markdown(f'<div class="comment-box danger">📌 최소 기준까지 {gap}건 부족 ({int(round(actual))}/{int(round(min_l))}건)</div>', unsafe_allow_html=True)

    # 개인레슨 3개월 비교 테이블
    if len(recent) > 1:
        rows = []
        items = [("전체고객수", False), ("개인레슨수", False), ("개인OT수", False),
                 ("개인레슨+OT", False), ("목표레슨수", False), ("개인출석률", "rate"), ("개인출석달성율", True)]
        for item_name, fmt_type in items:
            row = [item_name]
            for m in recent:
                ir = get_instructor_row(selected_inst, m)
                val = ir[item_name] if ir is not None else None
                if val is not None and not pd.isna(val):
                    if fmt_type == "rate":
                        row.append(f"{val:.2f}")
                    elif fmt_type:
                        row.append(f"{val:.1%}")
                    else:
                        row.append(str(int(round(val))))
                else:
                    row.append("-")
            rows.append(row)
        cols = ["항목"] + recent
        st.markdown(render_3month_table(rows, cols), unsafe_allow_html=True)

    # ===== 듀엣 섹션 =====
    section_header("👥 듀엣", "duet")
    dc1, dc2, dc3 = st.columns(3)
    with dc1:
        metric_card("듀엣회원수", fmt_num(r["듀엣회원수"]), color="blue")
    with dc2:
        metric_card("듀엣레슨수", fmt_num(r["듀엣레슨수"]), color="green")
    with dc3:
        metric_card("듀엣출석률", fmt_pct(r["듀엣출석률"]), color="orange")

    if len(recent) > 1:
        rows = []
        for item_name, is_pct in [("듀엣회원수", False), ("듀엣레슨수", False), ("듀엣출석률", True)]:
            row = [item_name]
            for m in recent:
                ir = get_instructor_row(selected_inst, m)
                val = ir[item_name] if ir is not None else None
                if val is not None and not pd.isna(val):
                    row.append(f"{val:.1%}" if is_pct else str(int(round(val))))
                else:
                    row.append("-")
            rows.append(row)
        st.markdown(render_3month_table(rows, ["항목"] + recent), unsafe_allow_html=True)

    # ===== 그룹 섹션 =====
    section_header("🏢 그룹", "group")
    gc1, gc2, gc3, gc4 = st.columns(4)
    with gc1:
        metric_card("그룹회원수", fmt_num(r["그룹회원수"]), color="blue")
    with gc2:
        metric_card("그룹수업수", fmt_num(r["그룹수업수"]), color="green")
    with gc3:
        metric_card("그룹출석수", fmt_num(r["그룹출석수"]), color="orange")
    with gc4:
        grp_color = "green" if r["그룹출석율"] and not pd.isna(r["그룹출석율"]) and r["그룹출석율"] >= 0.8 else "red"
        metric_card("그룹출석율", fmt_rate(r["그룹출석율"]), color=grp_color)

    if len(recent) > 1:
        rows = []
        for item_name, fmt_type in [("그룹회원수", False), ("그룹수업수", False), ("그룹출석수", False), ("그룹출석율", "rate")]:
            row = [item_name]
            for m in recent:
                ir = get_instructor_row(selected_inst, m)
                val = ir[item_name] if ir is not None else None
                if val is not None and not pd.isna(val):
                    if fmt_type == "rate":
                        row.append(f"{val:.2f}")
                    elif fmt_type:
                        row.append(f"{val:.1%}")
                    else:
                        row.append(str(int(round(val))))
                else:
                    row.append("-")
            rows.append(row)
        st.markdown(render_3month_table(rows, ["항목"] + recent), unsafe_allow_html=True)

    # ===== 체험 섹션 =====
    section_header("🌟 체험", "trial")
    tc1, tc2, tc3 = st.columns(3)
    with tc1:
        metric_card("체험수업수", fmt_num(r["체험수업수"]), color="orange")
    with tc2:
        metric_card("체험등록수", fmt_num(r["체험등록수"]), color="green")
    with tc3:
        trial_color = "green" if r["체험승률"] and not pd.isna(r["체험승률"]) and r["체험승률"] >= 0.5 else "red"
        metric_card("체험승률", fmt_pct(r["체험승률"]), color=trial_color)

    if len(recent) > 1:
        rows = []
        for item_name, is_pct in [("체험수업수", False), ("체험등록수", False), ("체험승률", True)]:
            row = [item_name]
            for m in recent:
                ir = get_instructor_row(selected_inst, m)
                val = ir[item_name] if ir is not None else None
                if val is not None and not pd.isna(val):
                    row.append(f"{val:.1%}" if is_pct else str(int(round(val))))
                else:
                    row.append("-")
            rows.append(row)
        st.markdown(render_3month_table(rows, ["항목"] + recent), unsafe_allow_html=True)

    # ===== 재등록 섹션 =====
    section_header("🔄 재등록", "renewal")
    rc1, rc2, rc3 = st.columns(3)
    with rc1:
        metric_card("재등예정수", fmt_num(r["재등예정수"]), color="orange")
    with rc2:
        metric_card("재등완료수", fmt_num(r["재등완료수"]), color="green")
    with rc3:
        re_color = "green" if r["재등록율"] and not pd.isna(r["재등록율"]) and r["재등록율"] >= 0.7 else "red"
        metric_card("재등록율", fmt_pct(r["재등록율"]), color=re_color)

    if len(recent) > 1:
        rows = []
        for item_name, is_pct in [("재등예정수", False), ("재등완료수", False), ("재등록율", True)]:
            row = [item_name]
            for m in recent:
                ir = get_instructor_row(selected_inst, m)
                val = ir[item_name] if ir is not None else None
                if val is not None and not pd.isna(val):
                    row.append(f"{val:.1%}" if is_pct else str(int(round(val))))
                else:
                    row.append("-")
            rows.append(row)
        st.markdown(render_3month_table(rows, ["항목"] + recent), unsafe_allow_html=True)

    # ===== 아카데미 섹션 =====
    section_header("🎓 아카데미", "")
    ac1, ac2, ac3, ac4, ac5 = st.columns(5)
    with ac1:
        metric_card("전체", fmt_num(r["아카데미수업수"]), color="gray")
    with ac2:
        metric_card("딥코칭", fmt_num(r.get("아카데미_딥코칭", 0)), color="blue")
    with ac3:
        metric_card("모의테스트", fmt_num(r.get("아카데미_모의테스트", 0)), color="orange")
    with ac4:
        metric_card("체험", fmt_num(r.get("아카데미_체험", 0)), color="green")
    with ac5:
        metric_card("그룹", fmt_num(r.get("아카데미_그룹", 0)), color="")

    if len(recent) > 1:
        rows = []
        for item_name in ["아카데미수업수", "아카데미_딥코칭", "아카데미_모의테스트", "아카데미_체험", "아카데미_그룹"]:
            row = [item_name.replace("아카데미_", "").replace("아카데미수업수", "전체")]
            for m in recent:
                ir = get_instructor_row(selected_inst, m)
                val = ir.get(item_name, 0) if ir is not None else 0
                if val is not None and not pd.isna(val):
                    row.append(str(int(round(val))))
                else:
                    row.append("0")
            rows.append(row)
        st.markdown(render_3month_table(rows, ["항목"] + recent), unsafe_allow_html=True)

    # ===== 종합 =====
    section_header("📊 종합 (스튜디오 평균 대비)", "")
    mdf = reports[current_month]
    avg_pa = mdf["개인출석률"].dropna().mean()
    avg_ga = mdf["그룹출석율"].dropna().mean()
    avg_rr = mdf["재등록율"].dropna().mean()
    avg_tr = mdf["체험승률"].dropna().mean()

    my_pa = r["개인출석률"]
    my_ga = r["그룹출석율"]
    my_rr = r["재등록율"]
    my_tr = r["체험승률"]

    def compare_color(my_val, avg_val):
        if my_val is None or pd.isna(my_val) or avg_val is None or pd.isna(avg_val):
            return "gray"
        return "green" if my_val >= avg_val else "red"

    zc1, zc2, zc3, zc4 = st.columns(4)
    with zc1:
        metric_card("개인출석률", fmt_rate(my_pa),
                    sub=f"평균 {fmt_rate(avg_pa)}",
                    color=compare_color(my_pa, avg_pa))
    with zc2:
        metric_card("그룹출석율", fmt_rate(my_ga),
                    sub=f"평균 {fmt_rate(avg_ga)}",
                    color=compare_color(my_ga, avg_ga))
    with zc3:
        metric_card("재등록률", fmt_pct(my_rr),
                    sub=f"평균 {fmt_pct(avg_rr)}",
                    color=compare_color(my_rr, avg_rr))
    with zc4:
        metric_card("체험승률", fmt_pct(my_tr),
                    sub=f"평균 {fmt_pct(avg_tr)}",
                    color=compare_color(my_tr, avg_tr))


# ============================================================
# PAGE: 월별 비교
# ============================================================
elif page == "📈 월별 비교":
    st.markdown("# 📈 월별 비교")

    if not months:
        st.info("데이터가 없습니다.")
        st.stop()

    # 연도 필터 (미래 확장용)
    all_years = sorted(set(["2025"]))  # 현재는 단일 연도, 나중에 데이터 기반으로 확장
    # 전체 월 선택 가능
    selected_months = st.multiselect("비교할 월 선택", months, default=months)

    if len(selected_months) < 1:
        st.info("1개 이상의 월을 선택해주세요.")
        st.stop()

    # ---- 스튜디오 전체 월별 추이 ----
    st.markdown("### 📊 스튜디오 전체 추이")
    summary_rows = []
    for m in selected_months:
        mdf = reports[m]
        summary_rows.append({
            "월": m,
            "총수업수": mdf["총수업수"].sum(),
            "개인레슨+OT": mdf["개인레슨+OT"].sum(),
            "듀엣레슨수": mdf["듀엣레슨수"].sum(),
            "그룹출석수": mdf["그룹출석수"].sum(),
            "평균개인출석률": mdf["개인출석률"].dropna().mean(),
            "평균재등록율": mdf["재등록율"].dropna().mean(),
            "평균체험승률": mdf["체험승률"].dropna().mean(),
            "평균그룹출석율": mdf["그룹출석율"].dropna().mean(),
            "강사수": len(mdf),
        })
    summary_df = pd.DataFrame(summary_rows)

    # 요약 카드
    if len(selected_months) >= 2:
        last = summary_df.iloc[-1]
        prev = summary_df.iloc[-2]
        c1, c2, c3, c4 = st.columns(4)
        with c1:
            d = last["총수업수"] - prev["총수업수"]
            metric_card("총수업수", fmt_num(last["총수업수"]),
                        sub=f"전월비 {'▲' if d>=0 else '▼'}{abs(d):.0f}건", color="blue")
        with c2:
            d = (last["평균개인출석률"] or 0) - (prev["평균개인출석률"] or 0)
            metric_card("평균개인출석률", fmt_rate(last["평균개인출석률"]),
                        sub=f"전월비 {'▲' if d>=0 else '▼'}{abs(d):.2f}", color="green")
        with c3:
            d = (last["평균재등록율"] or 0) - (prev["평균재등록율"] or 0)
            metric_card("평균재등록율", fmt_pct(last["평균재등록율"]),
                        sub=f"전월비 {'▲' if d>=0 else '▼'}{abs(d):.1%}p", color="orange")
        with c4:
            d = (last["평균그룹출석율"] or 0) - (prev["평균그룹출석율"] or 0)
            metric_card("평균그룹출석율", fmt_rate(last["평균그룹출석율"]),
                        sub=f"전월비 {'▲' if d>=0 else '▼'}{abs(d):.2f}", color="")

    # 차트
    st.markdown("---")
    col1, col2 = st.columns(2)
    with col1:
        fig = px.bar(summary_df, x="월", y=["총수업수", "개인레슨+OT", "듀엣레슨수", "그룹출석수"],
                     barmode="group", height=350,
                     color_discrete_sequence=["#2C3E50", "#3498db", "#9b59b6", "#27ae60"])
        fig.update_layout(legend_title="", margin=dict(l=0, r=0, t=10, b=0))
        st.plotly_chart(fig, use_container_width=True)
    with col2:
        fig2 = go.Figure()
        # 출석률 (소수점) — 막대
        for col_name, color in [("평균개인출석률", "#3498db"), ("평균그룹출석율", "#27ae60")]:
            fig2.add_trace(go.Bar(x=summary_df["월"], y=summary_df[col_name],
                                  name=col_name.replace("평균", ""),
                                  marker_color=color,
                                  text=[fmt_rate(v) for v in summary_df[col_name]],
                                  textposition="outside"))
        # 등록율 (%) — 꺽은선, 우측 Y축
        for col_name, color in [("평균재등록율", "#e74c3c"), ("평균체험승률", "#f39c12")]:
            fig2.add_trace(go.Scatter(x=summary_df["월"], y=summary_df[col_name],
                                      name=col_name.replace("평균", ""),
                                      mode="lines+markers+text", yaxis="y2",
                                      text=[fmt_pct(v) for v in summary_df[col_name]],
                                      textposition="top center",
                                      line=dict(color=color, width=2, dash="dot")))
        fig2.update_layout(height=350, margin=dict(l=0, r=0, t=30, b=0),
                          barmode="group",
                          yaxis=dict(title="출석률"),
                          yaxis2=dict(title="등록율(%)", overlaying="y", side="right", tickformat=".0%"),
                          legend_title="")
        st.plotly_chart(fig2, use_container_width=True)

    # ---- 강사별 월별 비교 테이블 ----
    st.markdown("### 📋 강사별 상세 비교")

    metric_choice = st.selectbox("지표 선택", [
        "총수업수", "개인레슨+OT", "개인출석률", "개인출석달성율", "듀엣레슨수", "듀엣출석률",
        "그룹출석수", "그룹출석율", "재등록율", "체험승률"
    ])

    # 소수점 표시 (% 아님): 개인출석률, 그룹출석율
    is_rate = metric_choice in ["개인출석률", "그룹출석율"]
    # 퍼센트 표시: 나머지 비율 지표
    is_pct = metric_choice in ["개인출석달성율", "듀엣출석률", "재등록율", "체험승률"]

    all_inst = sorted(set().union(*[set(reports[m]["강사"]) for m in selected_months]))
    rows = []
    for inst in all_inst:
        row = {"강사": inst}
        for m in selected_months:
            ir = get_instructor_row(inst, m)
            row[m] = ir[metric_choice] if ir is not None else None
        rows.append(row)

    comp_df = pd.DataFrame(rows)

    # 차트
    melt_df = comp_df.melt(id_vars="강사", var_name="월", value_name=metric_choice)
    fig = px.bar(melt_df, x="강사", y=metric_choice, color="월", barmode="group", height=400,
                 color_discrete_sequence=px.colors.qualitative.Set2)
    if is_pct:
        fig.update_layout(yaxis_tickformat=".0%")
    fig.update_layout(margin=dict(l=0, r=0, t=10, b=0), xaxis_tickangle=-45)
    st.plotly_chart(fig, use_container_width=True)

    # 테이블
    if is_pct:
        fmt_dict = {m: "{:.1%}" for m in selected_months}
    elif is_rate:
        fmt_dict = {m: "{:.2f}" for m in selected_months}
    else:
        fmt_dict = {m: "{:.0f}" for m in selected_months}
    st.dataframe(
        comp_df.style.format(fmt_dict, na_rep="-"),
        use_container_width=True, height=600
    )


# ============================================================
# PAGE: 팀별 비교
# ============================================================
elif page == "👥 팀별 비교":
    st.markdown("# 👥 팀별 비교")

    teams = config.get("teams", {})
    if not teams:
        st.warning("⚙️ 설정 페이지에서 팀 구성을 먼저 설정해주세요.")
        st.info("설정 → 팀 구성에서 강사를 팀에 배정하면 이 페이지에 팀별 비교가 표시됩니다.")
        st.stop()

    targets = config.get("targets", {})
    selected_month_team = st.selectbox("월 선택", months, index=len(months) - 1, key="team_month")
    df = reports[selected_month_team]

    # ---- 팀 간 핵심 4대 지표 비교 ----
    st.markdown("### 🏆 팀 간 핵심지표 비교")

    team_summary = []
    for team_name, members in teams.items():
        team_df = df[df["강사"].isin(members)]
        if team_df.empty:
            continue
        team_summary.append({
            "팀": team_name,
            "강사수": len(team_df),
            "개인출석률": team_df["개인출석률"].dropna().mean(),
            "그룹출석율": team_df["그룹출석율"].dropna().mean(),
            "재등록율": team_df["재등록율"].dropna().mean(),
            "체험승률": team_df["체험승률"].dropna().mean(),
            "총수업수": team_df["총수업수"].sum(),
            "개인레슨+OT": team_df["개인레슨+OT"].sum(),
        })

    if team_summary:
        ts_df = pd.DataFrame(team_summary)

        # 핵심 4대 지표 비교 차트
        fig = go.Figure()
        colors = {"개인출석률": "#3498db", "그룹출석율": "#27ae60",
                  "재등록율": "#e74c3c", "체험승률": "#f39c12"}
        for metric_name, color in colors.items():
            is_rate_metric = metric_name in ("개인출석률", "그룹출석율")
            fig.add_trace(go.Bar(
                x=ts_df["팀"], y=ts_df[metric_name],
                name=metric_name, marker_color=color,
                text=[fmt_rate(v) if is_rate_metric else fmt_pct(v) for v in ts_df[metric_name]],
                textposition="outside"
            ))
        fig.update_layout(barmode="group", height=350,
                         margin=dict(l=0, r=0, t=10, b=0),
                         legend_title="")
        st.plotly_chart(fig, use_container_width=True)

        # 팀별 요약 카드
        for _, ts in ts_df.iterrows():
            st.markdown(f"---")
            st.markdown(f"### {ts['팀']} ({int(ts['강사수'])}명)")

            # 핵심 4대 지표 카드
            k1, k2, k3, k4 = st.columns(4)
            with k1:
                key_metric_card("개인출석률", ts["개인출석률"],
                                target=targets.get("개인출석율_목표", 0.85), as_pct=False)
            with k2:
                key_metric_card("그룹출석율", ts["그룹출석율"],
                                target=targets.get("그룹출석율_목표", 0.80), as_pct=False)
            with k3:
                key_metric_card("재등록율", ts["재등록율"],
                                target=targets.get("재등록률_목표", 0.70))
            with k4:
                key_metric_card("체험승률", ts["체험승률"],
                                target=targets.get("체험등록율_목표", 0.50))

            # 세부 팀원 데이터 (expander)
            team_name = ts["팀"]
            members = teams[team_name]
            with st.expander(f"📋 {team_name} 팀원 세부 데이터 보기"):
                member_df = df[df["강사"].isin(members)].sort_values("강사")
                if member_df.empty:
                    st.info("해당 팀원 데이터가 없습니다.")
                else:
                    # 팀원별 핵심 4대 지표 테이블
                    header = "<tr><th>강사</th><th>전체고객수</th><th>개인레슨+OT</th><th>개인출석률</th><th>그룹출석수</th><th>그룹수업수</th><th>그룹출석율</th><th>재등예정</th><th>재등완료</th><th>재등록율</th><th>체험수업</th><th>체험등록</th><th>체험승률</th><th>총수업수</th></tr>"
                    body = ""
                    for _, mr in member_df.iterrows():
                        def val_class(val, target, as_pct=True):
                            if val is None or pd.isna(val):
                                return "", "-"
                            display = f"{val:.1%}" if as_pct else f"{val:.2f}"
                            if target and val >= target:
                                return "good-val", display
                            elif target and val >= target * 0.8:
                                return "warn-val", display
                            elif target:
                                return "bad-val", display
                            return "", display

                        pc, pv = val_class(mr["개인출석률"], targets.get("개인출석율_목표", 0.85), as_pct=False)
                        gc, gv = val_class(mr["그룹출석율"], targets.get("그룹출석율_목표", 0.80), as_pct=False)
                        rc, rv = val_class(mr["재등록율"], targets.get("재등록률_목표", 0.70))
                        tc, tv = val_class(mr["체험승률"], targets.get("체험등록율_목표", 0.50))

                        body += f"""<tr>
                            <td style="font-weight:600">{mr['강사']}</td>
                            <td>{fmt_num(mr['전체고객수'])}</td>
                            <td>{fmt_num(mr['개인레슨+OT'])}</td>
                            <td class="{pc}">{pv}</td>
                            <td>{fmt_num(mr['그룹출석수'])}</td>
                            <td>{fmt_num(mr['그룹수업수'])}</td>
                            <td class="{gc}">{gv}</td>
                            <td>{fmt_num(mr['재등예정수'])}</td>
                            <td>{fmt_num(mr['재등완료수'])}</td>
                            <td class="{rc}">{rv}</td>
                            <td>{fmt_num(mr['체험수업수'])}</td>
                            <td>{fmt_num(mr['체험등록수'])}</td>
                            <td class="{tc}">{tv}</td>
                            <td style="font-weight:600">{fmt_num(mr['총수업수'])}</td>
                        </tr>"""

                    st.markdown(f'<table class="member-table">{header}{body}</table>',
                                unsafe_allow_html=True)

                    # 팀원별 핵심지표 개별 차트
                    st.markdown("#### 팀원별 핵심지표")
                    mc1, mc2 = st.columns(2)
                    with mc1:
                        # 개인출석률
                        df_pa = member_df.dropna(subset=["개인출석률"]).sort_values("개인출석률", ascending=True)
                        if not df_pa.empty:
                            colors_pa = ["#e74c3c" if v < 1.0 else "#f39c12" if v < 1.5 else "#27ae60" for v in df_pa["개인출석률"]]
                            fig_pa = go.Figure(go.Bar(
                                x=df_pa["개인출석률"], y=df_pa["강사"], orientation="h",
                                marker_color=colors_pa,
                                text=[fmt_rate(v) for v in df_pa["개인출석률"]], textposition="outside",
                            ))
                            fig_pa.update_layout(height=max(180, len(df_pa) * 35), margin=dict(l=0, r=0, t=25, b=0), title="개인출석률")
                            st.plotly_chart(fig_pa, use_container_width=True)
                    with mc2:
                        # 그룹출석율
                        df_ga = member_df.dropna(subset=["그룹출석율"]).sort_values("그룹출석율", ascending=True)
                        if not df_ga.empty:
                            colors_ga = ["#e74c3c" if v < 2.0 else "#f39c12" if v < 3.0 else "#27ae60" if v < 4.0 else "#3498db" if v < 5.0 else "#8e44ad" for v in df_ga["그룹출석율"]]
                            fig_ga = go.Figure(go.Bar(
                                x=df_ga["그룹출석율"], y=df_ga["강사"], orientation="h",
                                marker_color=colors_ga,
                                text=[fmt_rate(v) for v in df_ga["그룹출석율"]], textposition="outside",
                            ))
                            fig_ga.update_layout(height=max(180, len(df_ga) * 35), margin=dict(l=0, r=0, t=25, b=0), title="그룹출석율")
                            st.plotly_chart(fig_ga, use_container_width=True)
                    mc3, mc4 = st.columns(2)
                    with mc3:
                        # 재등록률
                        df_rr = member_df.dropna(subset=["재등록율"]).sort_values("재등록율", ascending=True)
                        if not df_rr.empty:
                            fig_rr = go.Figure(go.Bar(
                                x=df_rr["재등록율"], y=df_rr["강사"], orientation="h",
                                marker_color=["#27ae60" if v >= 0.7 else "#f39c12" if v >= 0.5 else "#e74c3c" for v in df_rr["재등록율"]],
                                text=[fmt_pct(v) for v in df_rr["재등록율"]], textposition="outside",
                            ))
                            fig_rr.update_layout(height=max(180, len(df_rr) * 35), margin=dict(l=0, r=0, t=25, b=0),
                                                title="재등록률", xaxis_tickformat=".0%")
                            st.plotly_chart(fig_rr, use_container_width=True)
                    with mc4:
                        # 체험승률
                        df_tr = member_df.dropna(subset=["체험승률"]).sort_values("체험승률", ascending=True)
                        if not df_tr.empty:
                            fig_tr = go.Figure(go.Bar(
                                x=df_tr["체험승률"], y=df_tr["강사"], orientation="h",
                                marker_color=["#27ae60" if v >= 0.7 else "#f39c12" if v >= 0.5 else "#e74c3c" for v in df_tr["체험승률"]],
                                text=[fmt_pct(v) for v in df_tr["체험승률"]], textposition="outside",
                            ))
                            fig_tr.update_layout(height=max(180, len(df_tr) * 35), margin=dict(l=0, r=0, t=25, b=0),
                                                title="체험승률", xaxis_tickformat=".0%")
                            st.plotly_chart(fig_tr, use_container_width=True)

    # ---- 팀별 월별 추이 ----
    if len(months) > 1:
        st.markdown("---")
        st.markdown("### 📈 팀별 월별 추이")
        team_trend = []
        for m in months:
            mdf = reports[m]
            for team_name, members in teams.items():
                tdf = mdf[mdf["강사"].isin(members)]
                if tdf.empty:
                    continue
                team_trend.append({
                    "월": m, "팀": team_name,
                    "개인출석률": tdf["개인출석률"].dropna().mean(),
                    "그룹출석율": tdf["그룹출석율"].dropna().mean(),
                    "재등록율": tdf["재등록율"].dropna().mean(),
                    "체험승률": tdf["체험승률"].dropna().mean(),
                })
        if team_trend:
            tt_df = pd.DataFrame(team_trend)
            tc1, tc2 = st.columns(2)
            with tc1:
                fig = px.line(tt_df, x="월", y="개인출석률", color="팀",
                             markers=True, height=280)
                fig.update_layout(margin=dict(l=0, r=0, t=30, b=0),
                                  title="개인출석률 추이")
                st.plotly_chart(fig, use_container_width=True)
            with tc2:
                fig2 = px.line(tt_df, x="월", y="그룹출석율", color="팀",
                              markers=True, height=280)
                fig2.update_layout(margin=dict(l=0, r=0, t=30, b=0),
                                  title="그룹출석율 추이")
                st.plotly_chart(fig2, use_container_width=True)
            tc3, tc4 = st.columns(2)
            with tc3:
                fig3 = px.line(tt_df, x="월", y="재등록율", color="팀",
                              markers=True, height=280)
                fig3.update_layout(margin=dict(l=0, r=0, t=30, b=0),
                                  yaxis_tickformat=".0%", title="재등록율 추이")
                st.plotly_chart(fig3, use_container_width=True)
            with tc4:
                fig4 = px.line(tt_df, x="월", y="체험승률", color="팀",
                              markers=True, height=280)
                fig4.update_layout(margin=dict(l=0, r=0, t=30, b=0),
                                  yaxis_tickformat=".0%", title="체험승률 추이")
                st.plotly_chart(fig4, use_container_width=True)


# ============================================================
# PAGE: 설정
# ============================================================
elif page == "⚙️ 설정":
    st.markdown("# ⚙️ 설정")
    # 설정 페이지에서는 캐시 안 탄 config 직접 로드
    config = load_config()

    tab1, tab2, tab3, tab4, tab5, tab6 = st.tabs(
        ["👥 팀구성", "🎯 목표설정", "📁 데이터관리", "👩‍🏫 직원관리", "📄 상품리스트", "🔧 기타"]
    )

    # ── 탭1: 팀구성 ──
    with tab1:
        st.markdown("### 팀 구성")
        st.markdown("강사를 팀에 배정하세요. 팀명을 입력하고 해당 팀의 강사를 선택합니다.")

        teams = config.get("teams", {})
        # 재직 강사만 표시
        active_instructors = [nm for nm in config.get("instructors", [])
                              if config.get("staff", {}).get(nm, {}).get("status", "재직") == "재직"]
        num_teams = st.number_input("팀 수", min_value=1, max_value=10, value=max(len(teams), 2))

        new_teams = {}
        existing_team_names = list(teams.keys())
        all_assigned = set()

        for i in range(int(num_teams)):
            default_name = existing_team_names[i] if i < len(existing_team_names) else f"{chr(65 + i)}팀"
            col1, col2 = st.columns([1, 3])
            with col1:
                team_name = st.text_input(f"팀 {i + 1} 이름", value=default_name, key=f"team_name_{i}")
            with col2:
                default_members = teams.get(default_name, [])
                members = st.multiselect(f"팀 {i + 1} 강사", active_instructors,
                                        default=[m for m in default_members if m in active_instructors],
                                        key=f"team_members_{i}")
            if team_name and members:
                new_teams[team_name] = members
                all_assigned.update(members)

        unassigned = [i for i in active_instructors if i not in all_assigned]
        if unassigned:
            st.info(f"미배정 강사: {', '.join(unassigned)}")

        if st.button("💾 팀 구성 저장", type="primary"):
            config["teams"] = new_teams
            save_config(config)
            st.success("팀 구성이 저장되었습니다!")
            st.cache_data.clear()
            st.rerun()

    # ── 탭2: 목표설정 ──
    with tab2:
        st.markdown("### 목표 설정")
        targets = config.get("targets", {})

        st.markdown("#### 🏃 개인")
        c1, c2, c3 = st.columns(3)
        with c1:
            targets["최소레슨_배수"] = st.number_input("최소레슨 배수 (전체고객수 ×)", 0.5, 3.0,
                                                   targets.get("최소레슨_배수", 1.0), 0.1)
        with c2:
            targets["목표레슨_배수"] = st.number_input("목표레슨 배수 (전체고객수 ×)", 0.5, 3.0,
                                                   targets.get("목표레슨_배수", 1.5), 0.1)
        with c3:
            targets["최대레슨_배수"] = st.number_input("최대레슨 배수 (전체고객수 ×)", 0.5, 4.0,
                                                   targets.get("최대레슨_배수", 2.0), 0.1)
        c4, c5 = st.columns(2)
        with c4:
            targets["재등록률_목표"] = st.slider("재등록률 목표", 0, 100,
                                            int(targets.get("재등록률_목표", 0.70) * 100), 5, format="%d%%")
            targets["재등록률_목표"] = targets["재등록률_목표"] / 100.0
        with c5:
            targets["체험등록율_목표"] = st.slider("체험승률(체험등록율) 목표", 0, 100,
                                              int(targets.get("체험등록율_목표", 0.50) * 100), 5, format="%d%%")
            targets["체험등록율_목표"] = targets["체험등록율_목표"] / 100.0

        st.markdown("#### 🏢 그룹")
        targets["그룹출석율_목표"] = st.slider("그룹출석율 목표", 0.0, 6.0,
                                          float(targets.get("그룹출석율_목표", 3.0)), 0.1, format="%.1f")

        if st.button("💾 목표 설정 저장", type="primary"):
            config["targets"] = targets
            save_config(config)
            st.success("목표 설정이 저장되었습니다!")
            st.cache_data.clear()
            st.rerun()

    # ── 탭3: 데이터관리 ──
    with tab3:
        st.markdown("### 📤 엑셀 데이터 업로드")
        st.markdown("핏투데이케어에서 다운받은 `수업 및 정산관리` 파일을 업로드하세요.")

        upload_month = st.text_input("월 입력 (예: 2025년 4월)", placeholder="2025년 4월")
        uploaded_files = st.file_uploader(
            "엑셀 파일 선택 (여러 개 가능)", type=["xlsx", "xls"],
            accept_multiple_files=True, key="excel_upload"
        )

        if uploaded_files and upload_month:
            if st.button("📁 파일 저장 및 등록", type="primary"):
                data_dir = os.path.join(os.path.dirname(__file__), "data")
                os.makedirs(data_dir, exist_ok=True)
                saved_paths = []
                for i, uf in enumerate(uploaded_files, 1):
                    save_path = os.path.join(data_dir, f"{upload_month}_{i}.xlsx")
                    with open(save_path, "wb") as f:
                        f.write(uf.getbuffer())
                    saved_paths.append(save_path)
                config["data_files"][upload_month] = saved_paths
                save_config(config)
                st.cache_data.clear()
                st.success(f"{upload_month} 데이터 {len(saved_paths)}개 파일 저장 완료!")
                st.rerun()

        st.markdown("---")
        st.markdown("### 📋 현재 등록된 데이터")
        for month, files in sorted(config.get("data_files", {}).items()):
            with st.expander(f"📁 {month} ({len(files)}개 파일)"):
                for f in files:
                    exists = "✅" if os.path.exists(f) else "❌"
                    st.text(f"  {exists} {os.path.basename(f)}")
                if st.button(f"🗑️ {month} 데이터 삭제", key=f"del_{month}"):
                    del config["data_files"][month]
                    save_config(config)
                    st.cache_data.clear()
                    st.rerun()

    # ── 탭4: 직원관리 ──
    with tab4:
        st.markdown("### 👩‍🏫 직원관리 (이름 / 이메일 / 재직상태)")
        st.caption("신규 입사자는 하단 ➕ 버튼으로 추가, 퇴사자는 재직상태를 '퇴사'로 변경")

        # staff 데이터 초기화
        staff = config.get("staff", {})
        if not staff:
            for nm in config.get("instructors", []):
                staff[nm] = {"email": "", "status": "재직"}
            if config.get("instructor_emails"):
                for nm, em in config["instructor_emails"].items():
                    if nm in staff:
                        staff[nm]["email"] = em
                    else:
                        staff[nm] = {"email": em, "status": "재직"}
            config["staff"] = staff
            save_config(config)

        # 삭제 콜백
        def do_delete_staff(del_name):
            cfg = load_config()
            if del_name in cfg.get("staff", {}):
                del cfg["staff"][del_name]
                save_config(cfg)

        # 추가 콜백
        def do_add_staff():
            nm = st.session_state.get("new_staff_name", "").strip()
            em = st.session_state.get("new_staff_email", "").strip()
            if nm:
                cfg = load_config()
                if "staff" not in cfg:
                    cfg["staff"] = {}
                cfg["staff"][nm] = {"email": em, "status": "재직"}
                if nm not in cfg.get("instructors", []):
                    cfg.setdefault("instructors", []).append(nm)
                    cfg["instructors"].sort()
                save_config(cfg)
                st.session_state["new_staff_name"] = ""
                st.session_state["new_staff_email"] = ""

        # 다시 로드 (콜백 후 갱신)
        config = load_config()
        staff = config.get("staff", {})
        staff_list = sorted(staff.items())
        status_options = ["재직", "휴직", "퇴사"]

        if staff_list:
            hc1, hc2, hc3, hc4 = st.columns([2, 4, 2, 1])
            with hc1:
                st.markdown("**이름**")
            with hc2:
                st.markdown("**이메일**")
            with hc3:
                st.markdown("**재직상태**")

            for idx, (name, info) in enumerate(staff_list):
                sc1, sc2, sc3, sc4 = st.columns([2, 4, 2, 1])
                with sc1:
                    st.text(name)
                with sc2:
                    cur_email = info.get("email", "")
                    st.text_input("이메일", value=cur_email, key=f"staff_email_{idx}", label_visibility="collapsed")
                with sc3:
                    cur_status = info.get("status", "재직")
                    si = status_options.index(cur_status) if cur_status in status_options else 0
                    st.selectbox("상태", status_options, index=si, key=f"staff_status_{idx}", label_visibility="collapsed")
                with sc4:
                    st.button("🗑️", key=f"staff_del_{idx}",
                              on_click=do_delete_staff, args=(name,))

            active = sum(1 for _, v in staff_list if v.get("status") == "재직")
            leave = sum(1 for _, v in staff_list if v.get("status") == "휴직")
            retired = sum(1 for _, v in staff_list if v.get("status") == "퇴사")
            st.caption(f"총 {len(staff_list)}명 (재직 {active}명 / 휴직 {leave}명 / 퇴사 {retired}명)")
        else:
            st.info("등록된 직원이 없습니다.")

        # 저장 버튼
        def do_save_staff():
            cfg = load_config()
            staff_data = cfg.get("staff", {})
            for idx, (name, _) in enumerate(sorted(staff_data.items())):
                new_em = st.session_state.get(f"staff_email_{idx}", "")
                new_st = st.session_state.get(f"staff_status_{idx}", "재직")
                staff_data[name]["email"] = new_em
                staff_data[name]["status"] = new_st
            cfg["staff"] = staff_data
            save_config(cfg)

        st.button("💾 직원정보 저장", type="primary", on_click=do_save_staff)

        st.markdown("---")
        st.markdown("**➕ 새 직원 추가**")
        nf1, nf2 = st.columns([2, 4])
        with nf1:
            st.text_input("이름", key="new_staff_name", placeholder="이름 입력")
        with nf2:
            st.text_input("이메일", key="new_staff_email", placeholder="이메일 입력")
        st.button("➕ 추가", key="add_staff_btn", type="primary", on_click=do_add_staff)

        # ── SMTP 설정 ──
        st.markdown("---")
        smtp_config = config.get("smtp", {})
        smtp_status = f"✅ 설정됨 ({smtp_config.get('sender_name', '')} / {smtp_config.get('sender_email', '')})" if smtp_config.get("sender_email") else "⚠️ 미설정"
        with st.expander(f"📧 SMTP 설정 {smtp_status}"):
            st.caption("Gmail 사용 시: smtp.gmail.com / 587 / 앱 비밀번호 사용 ([Google 앱 비밀번호 생성](https://myaccount.google.com/apppasswords))")
            sender_name = st.text_input("발신자 이름", value=smtp_config.get("sender_name", "TRNT 필라테스"),
                                         placeholder="예: TRNT 필라테스")
            sm1, sm2 = st.columns(2)
            with sm1:
                smtp_server = st.text_input("SMTP 서버", value=smtp_config.get("server", "smtp.gmail.com"))
                sender_email = st.text_input("발신 이메일", value=smtp_config.get("sender_email", ""))
            with sm2:
                smtp_port = st.text_input("포트", value=str(smtp_config.get("port", "587")))
                sender_password = st.text_input("비밀번호 (앱 비밀번호)", value=smtp_config.get("sender_password", ""), type="password")

            if st.button("💾 SMTP 설정 저장"):
                config["smtp"] = {
                    "server": smtp_server,
                    "port": smtp_port,
                    "sender_email": sender_email,
                    "sender_password": sender_password,
                    "sender_name": sender_name,
                }
                save_config(config)
                st.success("SMTP 설정 저장 완료!")

        # ── 리포트 발송/다운로드 ──
        st.markdown("---")

        if months and reports:
            # 재직 강사만
            active_staff = [nm for nm, info in config.get("staff", {}).items() if info.get("status") == "재직"]
            all_instructors_set = sorted(set().union(*[set(reports[m]["강사"]) for m in months]))
            if active_staff:
                all_instructors_set = [i for i in all_instructors_set if i in active_staff]
            staff_emails = {nm: info.get("email", "") for nm, info in config.get("staff", {}).items()
                           if info.get("status") == "재직" and info.get("email")}
            registered = {inst: staff_emails[inst] for inst in all_instructors_set if inst in staff_emails}

            # 강사/월 선택을 위로
            rc1, rc2 = st.columns(2)
            with rc1:
                sel_inst_dl = st.selectbox("강사 선택", all_instructors_set, key="email_dl_inst")
            with rc2:
                sel_report_month = st.selectbox("리포트 기준 월", list(reversed(months)), key="report_month")

            # 해당월 포함 최근 3개월 계산
            mi = months.index(sel_report_month)
            report_recent = months[max(0, mi - 2):mi + 1]

            st.caption(f"리포트 포함 기간: {', '.join(report_recent)}")

            st.markdown("### 📤 리포트 발송 / 다운로드")
            dl1, dl2 = st.columns(2)
            with dl1:
                excel_buf = generate_instructor_report(sel_inst_dl, report_recent, reports, config)
                st.download_button(
                    f"📥 {sel_inst_dl} 리포트 다운로드",
                    data=excel_buf,
                    file_name=f"{sel_inst_dl}_레슨리포트_{report_recent[-1]}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            with dl2:
                zip_buf = generate_all_reports_zip(all_instructors_set, report_recent, reports, config)
                st.download_button(
                    "📦 전체 다운로드 (ZIP)",
                    data=zip_buf,
                    file_name=f"전체_레슨리포트_{report_recent[-1]}.zip",
                    mime="application/zip",
                )

            st.markdown("---")
            st.markdown("#### 📧 이메일 발송")
            if not smtp_config.get("sender_email"):
                st.warning("SMTP 설정을 먼저 완료해주세요.")
            elif not registered:
                st.warning("이메일이 등록된 재직 강사가 없습니다.")
            else:
                s1, s2 = st.columns(2)
                with s1:
                    inst_email = staff_emails.get(sel_inst_dl, "")
                    if inst_email:
                        if st.button(f"📧 {sel_inst_dl} 개별 발송", key="send_one_setting"):
                            with st.spinner("발송 중..."):
                                ok, msg_result = send_report_email(sel_inst_dl, inst_email, report_recent, reports, config)
                            if ok:
                                st.success(f"{sel_inst_dl} → {inst_email} 발송 완료!")
                            else:
                                st.error(f"실패: {msg_result}")
                    else:
                        st.info(f"{sel_inst_dl} 이메일 미등록")
                with s2:
                    if st.button(f"📧 전체 발송 ({len(registered)}명)", type="primary", key="send_all_setting"):
                        progress = st.progress(0)
                        results = []
                        for i, (inst, em) in enumerate(sorted(registered.items())):
                            ok, msg_result = send_report_email(inst, em, report_recent, reports, config)
                            results.append({"강사": inst, "이메일": em, "결과": "✅ 성공" if ok else f"❌ {msg_result}"})
                            progress.progress((i + 1) / len(registered))
                        st.dataframe(pd.DataFrame(results), use_container_width=True, hide_index=True)
                        success_count = sum(1 for rv in results if "성공" in rv["결과"])
                        st.success(f"{success_count}/{len(registered)}명 발송 완료")
        else:
            st.info("데이터가 없습니다. 데이터관리 탭에서 먼저 데이터를 업로드해주세요.")

        # 테스트 발송
        if smtp_config.get("sender_email"):
            with st.expander("🔧 테스트 발송"):
                test_email = st.text_input("테스트 수신 이메일", placeholder="테스트용 이메일 주소")
                if st.button("📧 테스트 메일 발송") and test_email:
                    try:
                        msg = MIMEMultipart()
                        msg["From"] = f"{smtp_config.get('sender_name', 'TRNT 필라테스')} <{smtp_config['sender_email']}>"
                        msg["To"] = test_email
                        msg["Subject"] = "[TRNT] 이메일 발송 테스트"
                        msg.attach(MIMEText("TRNT 레슨리포트 이메일 발송 테스트입니다.\n설정이 정상적으로 완료되었습니다!", "plain", "utf-8"))
                        with smtplib.SMTP(smtp_config.get("server", "smtp.gmail.com"), int(smtp_config.get("port", 587))) as server:
                            server.starttls()
                            server.login(smtp_config["sender_email"], smtp_config["sender_password"])
                            server.send_message(msg)
                        st.success("테스트 메일 발송 성공!")
                    except Exception as e:
                        st.error(f"발송 실패: {e}")

    # ── 탭5: 상품리스트 ──
    with tab5:
        st.markdown("### 📄 상품리스트 (수동 매핑)")
        st.caption("자동 분류되지 않는 이용권명을 직접 매핑합니다. 하단 ➕ 버튼으로 추가, 🗑️ 버튼으로 삭제")

        cat_options = ["개인", "듀엣", "그룹", "아카데미", "기타"]
        sub_options = {
            "개인": ["개인레슨", "개인OT"],
            "듀엣": ["듀엣레슨"],
            "그룹": ["그룹레슨"],
            "아카데미": ["딥코칭", "그룹"],
            "기타": ["기타"],
        }
        all_sub_options = ["개인레슨", "개인OT", "듀엣레슨", "그룹레슨", "딥코칭", "그룹", "기타"]

        manual_map = config.get("manual_product_map", {})
        prod_items = list(manual_map.items())

        if prod_items:
            # 헤더
            ph1, ph2, ph3, ph4 = st.columns([4, 2, 2, 1])
            with ph1:
                st.markdown("**이용권명**")
            with ph2:
                st.markdown("**대구분**")
            with ph3:
                st.markdown("**중구분**")

            for pidx, (pname, pcats) in enumerate(prod_items):
                pc1, pc2, pc3, pc4 = st.columns([4, 2, 2, 1])
                with pc1:
                    new_pname = st.text_input("이용권명", value=pname, key=f"prod_name_{pidx}", label_visibility="collapsed")
                with pc2:
                    cat1_idx = cat_options.index(pcats[0]) if pcats[0] in cat_options else 0
                    new_cat1 = st.selectbox("대구분", cat_options, index=cat1_idx, key=f"prod_cat1_{pidx}", label_visibility="collapsed")
                with pc3:
                    subs = sub_options.get(new_cat1, ["기타"])
                    cat2_idx = subs.index(pcats[1]) if pcats[1] in subs else 0
                    new_cat2 = st.selectbox("중구분", subs, index=cat2_idx, key=f"prod_cat2_{pidx}", label_visibility="collapsed")
                with pc4:
                    if st.button("🗑️", key=f"del_prod_{pidx}"):
                        del config["manual_product_map"][pname]
                        save_config(config)
                        st.cache_data.clear()
                        st.rerun()
                # 변경 감지
                if new_pname != pname or new_cat1 != pcats[0] or new_cat2 != pcats[1]:
                    if new_pname != pname and pname in config.get("manual_product_map", {}):
                        del config["manual_product_map"][pname]
                    config.setdefault("manual_product_map", {})[new_pname] = [new_cat1, new_cat2]
                    save_config(config)
                    st.cache_data.clear()

            st.caption(f"총 {len(prod_items)}개 수동 매핑")
        else:
            st.info("수동 매핑이 없습니다.")

        st.markdown("---")
        with st.form("add_prod_form"):
            st.markdown("**➕ 새 매핑 추가**")
            fp1, fp2, fp3 = st.columns([4, 2, 2])
            with fp1:
                new_prod_name = st.text_input("이용권명", placeholder="예: [개인] 개인 레슨 10회")
            with fp2:
                new_cat1 = st.selectbox("대구분", cat_options)
            with fp3:
                all_subs = ["개인레슨", "개인OT", "듀엣레슨", "그룹레슨", "딥코칭", "그룹", "모의테스트", "체험", "기타"]
                new_cat2 = st.selectbox("중구분", all_subs)
            prod_submitted = st.form_submit_button("➕ 추가", type="primary")
            if prod_submitted:
                if new_prod_name.strip():
                    if "manual_product_map" not in config:
                        config["manual_product_map"] = {}
                    config["manual_product_map"][new_prod_name.strip()] = [new_cat1, new_cat2]
                    save_config(config)
                    st.cache_data.clear()
                    st.success(f"'{new_prod_name.strip()}' 추가 완료!")
                    st.rerun()
                else:
                    st.warning("이용권명을 입력해주세요.")

        st.markdown("---")
        st.markdown("### 📋 전체 상품리스트 (현재 등록)")
        try:
            prod_map = load_product_list(config)
            prod_list_df = pd.DataFrame([
                {"이용권명": k, "TRNT대구분": v[0], "TRNT중구분": v[1]}
                for k, v in sorted(prod_map.items())
            ])
            with st.expander(f"전체 {len(prod_list_df)}개 상품 보기", expanded=False):
                cat_filter = st.selectbox("대구분 필터", ["전체"] + sorted(prod_list_df["TRNT대구분"].unique().tolist()), key="prod_filter")
                if cat_filter != "전체":
                    prod_list_df = prod_list_df[prod_list_df["TRNT대구분"] == cat_filter]
                st.dataframe(prod_list_df, use_container_width=True, height=400)
        except Exception as e:
            st.warning(f"상품리스트를 불러올 수 없습니다: {e}")

        st.markdown("---")
        st.markdown("### 📄 상품리스트 엑셀")
        st.text(f"📄 {config.get('product_list_path', '')}")
        prod_upload = st.file_uploader("새 상품리스트 업로드 (선택)", type=["xlsx"], key="prod_upload")
        if prod_upload:
            prod_path = os.path.join(os.path.dirname(__file__), "data", "TRNT_상품리스트.xlsx")
            os.makedirs(os.path.dirname(prod_path), exist_ok=True)
            with open(prod_path, "wb") as f:
                f.write(prod_upload.getbuffer())
            config["product_list_path"] = prod_path
            save_config(config)
            st.success("상품리스트 업데이트 완료!")

    # ── 탭6: 기타 (노션 연동) ──
    with tab6:
        st.markdown("### 🔗 노션 연동")
        new_token = st.text_input("노션 API 토큰", value=config.get("notion_api_token", ""), type="password")
        new_db_id = st.text_input("레슨리포트 DB ID", value=config.get("notion_db_id", ""))
        new_staff_db_id = st.text_input("인적사항 DB ID", value=config.get("notion_staff_db_id", ""),
                                         placeholder="인적사항 DB ID 입력")
        if st.button("💾 노션 설정 저장"):
            config["notion_api_token"] = new_token
            config["notion_db_id"] = new_db_id
            config["notion_staff_db_id"] = new_staff_db_id
            save_config(config)
            st.success("노션 설정 저장 완료!")

        st.markdown("---")
        st.markdown("### 🔒 대시보드 비밀번호 변경")
        new_pw = st.text_input("새 비밀번호", type="password", key="new_dashboard_pw")
        new_pw_confirm = st.text_input("비밀번호 확인", type="password", key="new_dashboard_pw_confirm")
        if st.button("💾 비밀번호 변경"):
            if not new_pw:
                st.warning("비밀번호를 입력해주세요.")
            elif new_pw != new_pw_confirm:
                st.error("비밀번호가 일치하지 않습니다.")
            else:
                config["dashboard_password"] = new_pw
                save_config(config)
                st.success("비밀번호가 변경되었습니다.")
